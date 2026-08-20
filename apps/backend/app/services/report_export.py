"""Rapor dışa aktarma — plan §4 "Faz 4", ADR §6.

`GET /api/v1/analyses/{id}/export?format=xlsx|json` gövdesini burası üretir.

## Sayılar METİN OLARAK YAZILMAZ

Bu modülün en önemli kuralı. `f"{value:.1f}".replace(".", ",")` gibi bir
Türkçe biçimleme cazip görünür ama hücreyi sayı olmaktan çıkarır: Excel'de
toplama, sıralama, grafik ve pivot çalışmaz, hücrenin köşesinde yeşil
"sayı olarak saklanmış metin" uyarısı çıkar. Adet/oran/güven alanları ham
`int` ve `float` olarak yazılır; biçimlendirmeyi Excel kendi yerelinde
yapar. Oranlar rapordaki değerin AYNISIDIR (0-100 aralığı), yeniden
hesaplanmaz — ADR §4: sayıları backend deterministik üretir, hiçbir
tüketici onları türetmez.

## Örnekler redakte edilmiş gelir

`redacted_examples` alanı `pipeline/aggregate.py`'de zaten maskelenmiş ve
kırpılmış olarak üretiliyor (ADR §9). Bu modül ham mesaja HİÇ erişmez;
elindeki tek kaynak rapor gövdesidir.
"""

from __future__ import annotations

from collections.abc import Sequence
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.schemas.common import to_iso_z
from app.schemas.report import AnalysisReport

#: xlsx MIME türü. Yanlış tür verilirse tarayıcı dosyayı zip sanıp açmaya
#: çalışır veya Excel "bozuk dosya" uyarısı gösterir.
XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

JSON_MEDIA_TYPE = "application/json"

#: Sayfa adları. Excel sayfa adı 31 karakterle sınırlı ve `[]:*?/\` kabul
#: etmiyor; kısa ve sabit tutuluyor.
SHEET_SUMMARY = "Özet"
SHEET_QUESTIONS = "Sorular"
SHEET_THEMES = "Temalar"

_HEADER_FONT = Font(bold=True)

#: Örnek mesajlar tek hücrede alt alta durur. Ayrı kolonlara yaymak, örnek
#: sayısı config'e bağlı olduğu için kolon şemasını değişken yapardı.
_EXAMPLE_SEPARATOR = "\n"

# Excel/LibreOffice, hücre metni bu karakterlerle başlarsa onu formül veya
# komut olarak yorumlayabilir. Kullanıcı denetimli kaynak adı, kolon/değer
# eşlemeleri ve LLM tarafından üretilen etiketler export'a taşındığı için
# koruma yalnız B4 metadata'sına değil bütün metin hücrelerine uygulanır.
_FORMULA_PREFIXES = ("=", "+", "-", "@")
_FORMULA_LEADING_WHITESPACE = " \t\r\n"

CellValue = str | int | float | None


def _safe_xlsx_value(value: CellValue) -> CellValue:
    """Kullanıcı/model metninin spreadsheet formülü olmasını engelle."""
    if isinstance(value, str) and value.lstrip(_FORMULA_LEADING_WHITESPACE).startswith(
        _FORMULA_PREFIXES
    ):
        return f"'{value}"
    return value


def _append_row(sheet: Worksheet, values: Sequence[CellValue]) -> None:
    """Metni sanitize ederken sayısal hücre tiplerini aynen koru."""
    sheet.append([_safe_xlsx_value(value) for value in values])


def content_disposition(analysis_id: str, extension: str) -> str:
    """`Content-Disposition` başlığı üretir.

    Dosya adı KULLANICININ yüklediği addan türetilmez; analiz kimliğinden
    üretilir ve ASCII kalır. İki sebebi var: (1) frontend'in
    `filenameFromContentDisposition` fonksiyonu yalnızca düz
    `filename="..."` biçimini ayrıştırıyor, RFC 5987'nin `filename*`
    kodlamasını değil — Türkçe karakterli bir dosya adı orada bozulurdu;
    (2) kullanıcı dosya adı denetimli değil, başlığa tırnak veya satır sonu
    enjekte edilebilirdi.
    """
    return f'attachment; filename="analiz-{analysis_id}.{extension}"'


def _write_header(sheet: Worksheet, headers: list[str]) -> None:
    _append_row(sheet, headers)
    for cell in sheet[1]:
        cell.font = _HEADER_FONT
    sheet.freeze_panes = "A2"


def _autosize(sheet: Worksheet, widths: list[int]) -> None:
    """Kolon genişliklerini sabitler.

    Otomatik ölçüm yapılmıyor: tüm satırları gezmek 100.000 kayıtlık bir
    raporda pahalı ve kazancı kozmetik.
    """
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def _build_summary_sheet(sheet: Worksheet, report: AnalysisReport) -> None:
    """Analizin künyesi: kaynak, ön işleme sayaçları, model ve maliyet.

    Sayaçlar ham sayı olarak yazılır; "Analiz edilen" ile "Elenen"in
    toplamı Excel'de doğrudan `SUM` ile doğrulanabilsin.
    """
    _write_header(sheet, ["Alan", "Değer"])

    source = report.source_summary
    pre = report.preprocessing_summary
    serialized_filters = "; ".join(
        f"{row_filter.column} = {' | '.join(row_filter.allowed_values)}"
        for row_filter in source.row_filters
    )

    rows: list[tuple[str, str | int | float]] = [
        ("Analiz kimliği", str(report.analysis_id)),
        ("Oluşturulma", to_iso_z(report.generated_at)),
        ("Dosya", source.filename),
        ("Sayfa", source.sheet_name),
        ("Kolon", source.text_column),
        (
            "Analiz biçimi",
            "Session bağlamlı kullanıcı turn'leri"
            if source.analysis_mode == "contextual_user_turns"
            else "Bağımsız mesajlar",
        ),
        ("Satır filtreleri", serialized_filters or "Yok"),
        ("Toplam satır", source.total_rows),
        ("Analiz edilen kayıt", pre.analyzed_count),
        ("Bağlam adayı (hedef dışı)", pre.context_only_count),
        ("Elenen kayıt", pre.discarded_count),
        ("Tekrar eden kayıt", pre.duplicate_count),
        ("Benzersiz kayıt", pre.unique_count),
        ("PII maskelenen kayıt", pre.redacted_count),
        ("Model", report.model),
        ("Prompt sürümü", report.prompt_version),
        ("Prompt özeti", report.prompt_hash),
        ("Prompt token", report.token_usage.prompt_tokens),
        ("Yanıt token", report.token_usage.completion_tokens),
        ("Cache'den okunan token", report.token_usage.cached_tokens),
        ("Cache'e yazılan token", report.token_usage.cache_write_tokens),
        ("Toplam token", report.token_usage.total_tokens),
        ("Maliyet (USD)", report.estimated_cost_usd),
        (
            "Maliyet kaynağı",
            "OpenRouter usage.cost"
            if report.cost_source == "provider"
            else "Fiyat snapshot hesabı",
        ),
    ]
    if source.conversation_config is not None:
        config = source.conversation_config
        rows.extend(
            [
                ("Session kolonu", config.session_id_column),
                ("Mesaj sıra kolonu", config.message_order_column),
                ("Rol kolonu", config.role_column),
                ("Mesaj türü kolonu", config.message_type_column),
                ("Kullanıcı rol değerleri", " | ".join(config.user_role_values)),
                ("Bot rol değerleri", " | ".join(config.assistant_role_values)),
                ("Hedef mesaj türleri", " | ".join(config.target_message_types)),
                ("Bağlam mesaj türleri", " | ".join(config.context_message_types)),
                ("Azami geçmiş turn", config.max_context_turns),
                ("Azami bağlam token", config.max_context_tokens),
            ]
        )
    if report.pricing_snapshot is not None:
        snapshot = report.pricing_snapshot
        rows.extend(
            [
                ("Fiyat kataloğu", snapshot.source),
                ("Girdi fiyatı (USD / 1M)", snapshot.input_cost_per_million),
                ("Çıktı fiyatı (USD / 1M)", snapshot.output_cost_per_million),
                (
                    "Cache okuma fiyatı (USD / 1M)",
                    snapshot.cache_read_cost_per_million
                    if snapshot.cache_read_cost_per_million is not None
                    else "Yok",
                ),
                (
                    "Cache yazma fiyatı (USD / 1M)",
                    snapshot.cache_write_cost_per_million
                    if snapshot.cache_write_cost_per_million is not None
                    else "Yok",
                ),
            ]
        )
    for label, value in rows:
        _append_row(sheet, [label, value])

    _append_row(sheet, [])
    _append_row(sheet, ["Yönetici özeti", report.executive_summary])
    sheet.cell(row=sheet.max_row, column=2).alignment = Alignment(wrap_text=True, vertical="top")

    if report.warnings:
        _append_row(sheet, [])
        _append_row(sheet, ["Uyarı kodu", "Açıklama"])
        for warning in report.warnings:
            _append_row(sheet, [warning.code, warning.message])

    _autosize(sheet, [26, 80])


def _build_questions_sheet(sheet: Worksheet, report: AnalysisReport) -> None:
    _write_header(
        sheet,
        ["Kimlik", "Soru", "Adet", "Oran (%)", "Örnek mesajlar (redakte)"],
    )
    for question in report.top_questions:
        _append_row(
            sheet,
            [
                question.id,
                question.canonical_question,
                # ---- ham sayılar: `int` ve `float`, dize DEĞİL ----
                question.count,
                question.percentage,
                _EXAMPLE_SEPARATOR.join(question.redacted_examples),
            ],
        )
        sheet.cell(row=sheet.max_row, column=5).alignment = Alignment(
            wrap_text=True, vertical="top"
        )

    _autosize(sheet, [16, 60, 10, 12, 70])


def _build_themes_sheet(sheet: Worksheet, report: AnalysisReport) -> None:
    _write_header(sheet, ["Kimlik", "Tema", "Adet", "Oran (%)", "İlgili soru kimlikleri"])
    for theme in report.themes:
        _append_row(
            sheet,
            [
                theme.id,
                theme.name,
                theme.count,
                theme.percentage,
                # Plan §1.2: bu liste top_n kırpmasından SONRA filtrelenmiş
                # hâliyle gelir; tema `count`'u ise kırpmadan etkilenmez.
                ", ".join(theme.related_question_ids),
            ],
        )

    _autosize(sheet, [16, 40, 10, 12, 50])


def build_xlsx(report: AnalysisReport) -> bytes:
    """Raporu üç sayfalı bir `.xlsx` gövdesine çevirir.

    SENKRON ve CPU yoğun: çağıran taraf bunu bir thread'e taşımalı, aksi
    hâlde büyük bir raporun yazımı event loop'u bloklar.
    """
    workbook = Workbook()
    # `Workbook()` "Sheet" adlı boş bir sayfayla geliyor; onu yeniden
    # adlandırıp kullanıyoruz, yoksa dosyada bir boş sayfa kalırdı.
    summary = workbook.active
    assert summary is not None
    summary.title = SHEET_SUMMARY

    _build_summary_sheet(summary, report)
    _build_questions_sheet(workbook.create_sheet(SHEET_QUESTIONS), report)
    _build_themes_sheet(workbook.create_sheet(SHEET_THEMES), report)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
