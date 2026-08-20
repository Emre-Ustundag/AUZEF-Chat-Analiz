"""Güvenli `.xlsx` doğrulama ve sayfa/kolon profilleme.

ADR §5 "Aşama A", §9 ve plan §3.2 (f)-(g).

Bu modül bilinçli olarak SAF tutulmuştur: veritabanı, S3 veya Celery bilmez,
yalnızca yerel bir dosya yolu alır. Faz 1'in tek gerçekten zor kodu burasıdır
ve altyapı olmadan test edilebilmesi gerekir.

Tehdit modeli — buraya gelen dosya GÜVENİLMEYEN kullanıcı girdisidir:

* ZIP bombası: birkaç yüz KB'lık bir zip gigabaytlarca açılabilir. openpyxl'e
  doğrudan verilirse worker'ın belleğini tüketir. Savunma iki katmanlı:
  önce zip dizinindeki beyan edilen boyutlar, sonra GERÇEK açılmış bayt
  sayısını sayan akışlı doğrulama. Beyan edilen boyutlar yalan söyleyebilir,
  bu yüzden ikincisi zorunludur.
* Makrolu dosya: `.xlsm` içeriği `.xlsx` adıyla gönderilebilir; ayırt edici
  işaret zip içindeki `xl/vbaProject.bin` üyesidir, dosya uzantısı değil.
* Şifreli dosya: OOXML değil, OLE2/CFB kabıdır ve `PK` ile başlamaz.
* Zip slip: üye adlarında `..` veya mutlak yol. Biz üyeleri diske
  açmıyoruz ama kontrol yine de yapılır — ileride biri açmaya kalkarsa
  savunma yerinde olsun.
"""

from __future__ import annotations

import zipfile
from collections.abc import Iterator, Sequence
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.core.config import MAX_ROWS, Settings
from app.core.logging import get_logger
from app.services.redaction import sanitize_sample

logger = get_logger(__name__)

#: OOXML/zip dosyalarının ilk dört baytı.
ZIP_MAGIC = b"PK\x03\x04"
#: Boş bir zip arşivinin imzası — geçerli bir xlsx asla boş olmaz.
ZIP_EMPTY_MAGIC = b"PK\x05\x06"
#: OLE2/CFB kabı: eski `.xls` VE parola korumalı `.xlsx` bu imzayı taşır.
OLE2_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"

#: Zip içinde bulunması zorunlu OOXML işaretleri.
REQUIRED_MEMBERS = ("[Content_Types].xml",)
WORKBOOK_MEMBERS = ("xl/workbook.xml", "xl/workbook.bin")

#: Makro göstergesi.
MACRO_MEMBERS = ("xl/vbaProject.bin",)
#: Parola korumalı OOXML paketinin OLE2 kabı içindeki adı.
ENCRYPTED_MEMBERS = ("EncryptedPackage", "EncryptionInfo")

#: Sıkıştırma oranı kontrolünün altına inilmeyecek üye boyutu. Küçük XML
#: parçalarında oran doğal olarak çok yüksek çıkar (100 baytlık dosya 20 bayta
#: iner) ve anlamlı bir sinyal taşımaz.
RATIO_CHECK_MIN_BYTES = 1 * 1024 * 1024

#: Akışlı doğrulamada kullanılan okuma bloğu.
_CHUNK_SIZE = 1024 * 1024

#: `unique_count` için bellek tavanı. Bunun üstünde benzersiz değer sayımı
#: alt sınır olarak raporlanır; 100 binlik bir kolonda tam sayım uğruna
#: worker'ın belleğini riske atmaya değmez.
_UNIQUE_TRACKING_LIMIT = 200_000


class XlsxRejectedError(Exception):
    """Dosya güvenlik veya biçim kontrolünden geçemedi.

    Çağıran bunu `UPLOAD_CORRUPT_OR_ENCRYPTED` (422) hatasına çevirir.
    `reason` yalnızca log ve test içindir; kullanıcıya frontend'in kendi
    Türkçe metni gösterilir.
    """

    def __init__(self, reason: str) -> None:
        super().__init__(reason)
        self.reason = reason


@dataclass
class ColumnStats:
    """Profil çıkarılırken tutulan ara sayaçlar."""

    name: str
    index: int
    non_empty_count: int = 0
    empty_count: int = 0
    total_length: int = 0
    string_count: int = 0
    unique_values: set[str] = field(default_factory=set)
    unique_overflow: bool = False
    samples: list[str] = field(default_factory=list)

    def observe(self, value: Any, *, settings: Settings) -> None:
        if value is None:
            self.empty_count += 1
            return

        text = str(value).strip() if not isinstance(value, str) else value.strip()
        if not text:
            self.empty_count += 1
            return

        self.non_empty_count += 1
        self.total_length += len(text)
        if isinstance(value, str):
            self.string_count += 1

        if not self.unique_overflow:
            if len(self.unique_values) < _UNIQUE_TRACKING_LIMIT:
                self.unique_values.add(text)
            else:
                self.unique_overflow = True

        # Örnekler yalnızca metin hücrelerden ve birbirinden farklı olacak
        # şekilde toplanır; üç kez aynı değeri göstermek kullanıcıya kolonun
        # ne içerdiğini anlatmaz.
        if len(self.samples) < settings.sample_values_per_column:
            sample = sanitize_sample(text, settings.sample_value_max_length)
            if sample and sample not in self.samples:
                self.samples.append(sample)

    @property
    def avg_length(self) -> float:
        if self.non_empty_count == 0:
            return 0.0
        return round(self.total_length / self.non_empty_count, 2)

    @property
    def unique_count(self) -> int:
        return len(self.unique_values)

    @property
    def is_likely_text(self) -> bool:
        """Kolonun serbest metin olup olmadığına dair BASİT bir sezgisel.

        Yalnızca bir öneridir (plan §3.2 g: "Kesin olması gerekmiyor,
        kullanıcı zaten seçiyor"). İki koşul aranır:

        * Değerlerin büyük çoğunluğu gerçekten metin tipinde olmalı. Bu,
          `data_only=True` ile tarih ve sayı kolonlarının `datetime`/`int`
          olarak gelmesi sayesinde tarih kolonlarını eler — ki bunların
          ortalama uzunluğu da metin sınırının üstünde çıkabiliyor.
        * Ortalama uzunluk bir eşiğin üstünde olmalı; "web"/"mobil" gibi
          kategorik metin kolonlarını eler.
        """
        if self.non_empty_count == 0:
            return False
        string_ratio = self.string_count / self.non_empty_count
        return string_ratio >= 0.8 and self.avg_length >= 15


def _detect_signature(path: Path) -> None:
    """Magic bytes kontrolü. ZIP değilse dosya OOXML olamaz."""
    with path.open("rb") as handle:
        header = handle.read(8)

    if header.startswith(OLE2_MAGIC):
        # Parola korumalı xlsx dosyaları OOXML zip'i bir OLE2 kabına sarar.
        # Eski `.xls` de aynı imzayı taşır; ikisi de reddedilir (ADR §9).
        raise XlsxRejectedError("encrypted_or_legacy_ole2_container")

    if header.startswith(ZIP_EMPTY_MAGIC):
        raise XlsxRejectedError("empty_zip_archive")

    if not header.startswith(ZIP_MAGIC):
        raise XlsxRejectedError("not_a_zip_archive")


def _check_members(archive: zipfile.ZipFile) -> None:
    """OOXML yapısı, makro ve şifreleme işaretlerini denetler."""
    names = set(archive.namelist())

    for member in ENCRYPTED_MEMBERS:
        if member in names:
            raise XlsxRejectedError("encrypted_ooxml_package")

    for member in MACRO_MEMBERS:
        if member in names:
            # Uzantı `.xlsx` olsa bile içerik makrolu bir çalışma kitabıdır.
            raise XlsxRejectedError("macro_enabled_workbook")

    for member in REQUIRED_MEMBERS:
        if member not in names:
            raise XlsxRejectedError("missing_ooxml_content_types")

    if not any(member in names for member in WORKBOOK_MEMBERS):
        raise XlsxRejectedError("missing_workbook_part")

    for name in names:
        # Zip slip savunması: bu modül üyeleri diske açmıyor ama kontrolü
        # burada tutmak, ileride açan bir kodun güvenli başlamasını sağlar.
        if name.startswith("/") or ".." in Path(name).parts:
            raise XlsxRejectedError("unsafe_member_path")


def _check_declared_sizes(archive: zipfile.ZipFile, settings: Settings) -> None:
    """ZIP bomba savunması, birinci katman: dizinde BEYAN EDİLEN boyutlar.

    Ucuzdur ve dosyayı hiç açmadan çalışır. Tek başına yeterli DEĞİLDİR —
    beyan edilen boyutlar saldırgan tarafından yazılır.
    """
    total_declared = 0
    for info in archive.infolist():
        total_declared += info.file_size
        if total_declared > settings.max_uncompressed_bytes:
            raise XlsxRejectedError("declared_uncompressed_size_exceeds_limit")

        if info.file_size >= RATIO_CHECK_MIN_BYTES and info.compress_size > 0:
            ratio = info.file_size / info.compress_size
            if ratio > settings.max_compression_ratio:
                raise XlsxRejectedError("compression_ratio_exceeds_limit")


def _check_actual_sizes(archive: zipfile.ZipFile, settings: Settings) -> int:
    """ZIP bomba savunması, ikinci katman: GERÇEKTEN açılan bayt sayısı.

    Her üyeyi bloklar hâlinde açar ve koşan bir toplam tutar; sınır aşılınca
    okuma anında kesilir. Beyan edilen boyutlara güvenilemeyeceği için asıl
    savunma budur: dizinde "1 KB" yazan bir üye 10 GB açılıyorsa ancak burada
    yakalanır.

    Bedeli dosyanın bir kez fazladan açılmasıdır. Bu, openpyxl'in aynı veriyi
    kontrolsüz açmasına izin vermenin yanında ucuz kalır.
    """
    total_actual = 0
    for info in archive.infolist():
        if info.is_dir():
            continue
        try:
            with archive.open(info) as member:
                while chunk := member.read(_CHUNK_SIZE):
                    total_actual += len(chunk)
                    if total_actual > settings.max_uncompressed_bytes:
                        raise XlsxRejectedError("actual_uncompressed_size_exceeds_limit")
        except (zipfile.BadZipFile, EOFError, OSError) as exc:
            raise XlsxRejectedError("member_decompression_failed") from exc
    return total_actual


def validate_xlsx(path: Path, settings: Settings) -> int:
    """Dosyayı güvenlik ve biçim açısından doğrular.

    Başarılıysa açılmış toplam bayt sayısını döndürür; aksi hâlde
    `XlsxRejectedError` fırlatır. Profil çıkarmadan ÖNCE çağrılmalıdır.
    """
    if not path.exists() or path.stat().st_size == 0:
        raise XlsxRejectedError("empty_file")

    _detect_signature(path)

    try:
        with zipfile.ZipFile(path) as archive:
            # `testzip` CRC uyuşmazlıklarını yakalar ama bomba kontrolünden
            # ÖNCE çağrılamaz: kendisi tüm arşivi açar.
            _check_members(archive)
            _check_declared_sizes(archive, settings)
            uncompressed = _check_actual_sizes(archive, settings)

            corrupt_member = archive.testzip()
            if corrupt_member is not None:
                raise XlsxRejectedError("crc_mismatch")
    except zipfile.BadZipFile as exc:
        raise XlsxRejectedError("corrupt_zip_archive") from exc
    except XlsxRejectedError:
        raise
    except OSError as exc:
        raise XlsxRejectedError("unreadable_file") from exc

    return uncompressed


def _column_name(raw: Any, index: int) -> str:
    """Başlık hücresinden kolon adı üretir.

    Boş başlıklar için Excel'in kullanıcıya gösterdiğine yakın bir ad
    üretiliyor; `None` döndürmek frontend'in `name: z.string()` sözleşmesini
    bozardı.
    """
    if raw is None:
        return f"Kolon {index + 1}"
    text = str(raw).strip()
    return text or f"Kolon {index + 1}"


def profile_xlsx(path: Path, settings: Settings) -> dict[str, Any]:
    """Doğrulanmış bir `.xlsx`'ten sayfa ve kolon profili çıkarır.

    `read_only=True, data_only=True` ile satır satır okunur: ADR §2'nin
    öngördüğü gibi 130 MB'lık bir dosya belleğe tamamen alınmaz.

    Dönen sözlük `UploadProfile` şemasıyla birebir uyumludur.
    """
    try:
        workbook = load_workbook(
            path,
            read_only=True,
            data_only=True,
            keep_links=False,
        )
    except XlsxRejectedError:
        raise
    except Exception as exc:  # openpyxl çok çeşitli istisna tipleri atıyor
        raise XlsxRejectedError("workbook_parse_failed") from exc

    sheets: list[dict[str, Any]] = []
    total_row_count = 0

    try:
        for worksheet in workbook.worksheets:
            sheet_profile, row_count = _profile_sheet(worksheet, settings)
            sheets.append(sheet_profile)
            total_row_count += row_count
    finally:
        workbook.close()

    return {
        "sheets": sheets,
        "total_row_count": total_row_count,
        # ADR §9 / plan §3.2 (g): sınır aşılırsa iş BAŞARISIZ OLMAZ,
        # yalnızca işaretlenir ve arayüz kullanıcıyı uyarır.
        "exceeds_row_limit": total_row_count > MAX_ROWS,
    }


def _profile_sheet(worksheet: Any, settings: Settings) -> tuple[dict[str, Any], int]:
    columns: list[ColumnStats] = []
    row_count = 0
    header_seen = False

    for row in worksheet.iter_rows(values_only=True):
        if not header_seen:
            # İlk satır başlık kabul edilir. Kolon SAYISI da buradan gelir;
            # sonraki satırlarda fazladan hücre varsa yeni kolon açılır.
            for index, value in enumerate(row):
                columns.append(ColumnStats(name=_column_name(value, index), index=index))
            header_seen = True
            continue

        if row_count >= settings.profile_max_scan_rows:
            # Taranan satır sayısına son çare tavanı — `break`, `continue`
            # DEĞİL.
            #
            # Eskiden `continue` idi ve sayacı artırmaya devam ediyordu
            # ("`exceeds_row_limit` doğru hesaplansın" diye). Sonuç: kolon
            # istatistikleri ilk N satırı, `row_count` ise dosyanın tamamını
            # anlatıyordu ve `UploadProfile`'ın değişmezi
            # (`non_empty_count + empty_count == row_count`) İHLAL EDİLİYORDU.
            # Pydantic doğrulaması patlıyor, kullanıcı `INTERNAL_ERROR`
            # alıyordu — yani tavanı aşan HER dosya profillenemiyordu. Yük
            # testi (ADR §10 risk 1) bunu 2,8 M satırlık gerçek bir dosyada
            # yakaladı; ADR-0002 #13 ise "dosya tam profillenir, reddedilmez"
            # diyor.
            #
            # `break` ile profil, GERÇEKTEN taranmış satırları anlatır ve
            # değişmez yapı gereği doğru kalır. Tavan `MAX_ROWS`'un çok
            # üstünde tutulduğu için (config doğrulaması bunu zorunlu kılıyor)
            # `exceeds_row_limit` yine doğru işaretlenir.
            break

        # Tamamen boş satırlar Excel'in "kullanılmış aralık" şişmesi yüzünden
        # sıkça görülür; satır sayısına katılmazlar.
        if all(value is None or (isinstance(value, str) and not value.strip()) for value in row):
            continue

        row_count += 1

        for index, value in enumerate(row):
            if index >= len(columns):
                columns.append(ColumnStats(name=f"Kolon {index + 1}", index=index))
            columns[index].observe(value, settings=settings)

    return (
        {
            "name": worksheet.title,
            "row_count": row_count,
            "column_count": len(columns),
            "columns": [
                {
                    "name": column.name,
                    "index": column.index,
                    "non_empty_count": column.non_empty_count,
                    "empty_count": column.empty_count,
                    "unique_count": column.unique_count,
                    "avg_length": column.avg_length,
                    "is_likely_text": column.is_likely_text,
                    "sample_values": column.samples,
                }
                for column in columns
            ],
        },
        row_count,
    )


class SheetOrColumnNotFoundError(Exception):
    """Seçilen sayfa veya kolon dosyada yok.

    Çağıran bunu `SHEET_OR_COLUMN_NOT_FOUND` (422) hatasına çevirir. Kullanıcı
    kolon seçimini yaptıktan sonra dosya değişmiş olabilir ya da istek elle
    hazırlanmış olabilir; her iki durumda da iş burada durur.
    """

    def __init__(self, reason: str) -> None:
        super().__init__(reason)
        self.reason = reason


def iter_row_values(
    path: Path,
    sheet_name: str,
    columns: Sequence[str],
) -> Iterator[tuple[str | None, ...]]:
    """Seçilen sayfanın seçilen KOLONLARINI satır satır, istenen sırada üretir.

    ADR §5 Aşama B madde 1. Faz 1'in profilleme kodu gibi
    `read_only=True, data_only=True` kullanır: 130 MB'lık bir dosyanın tamamı
    belleğe ALINMAZ; yalnızca istenen kolonlar üretilir.

    Kolonlar, profil çıkarılırken kullanılan mantığın AYNISIYLA bulunur
    (`_column_name`): ilk satır başlıktır ve boş başlıklar `Kolon N` olur.
    İki yerde farklı isimlendirme kullanmak, kullanıcının arayüzde seçtiği
    kolon adının burada bulunamamasına yol açardı.

    Boş hücreler `None` olarak ÜRETİLİR, atlanmaz: `discarded_count` ve
    `total_rows` sayımlarının doğru olması için toplam satır sayısı gerekli.

    Çok kolonlu biçim `CHATBOT_LOG` ön ayarı için eklendi: rol, oturum ve
    zaman kolonları metin kolonuyla AYNI geçişte okunur; dosyayı kolon başına
    yeniden taramak 130 MB'lık gerçek dökümde okuma süresini katlardı.
    """
    try:
        workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    except Exception as exc:
        raise XlsxRejectedError("workbook_parse_failed") from exc

    try:
        if sheet_name not in workbook.sheetnames:
            raise SheetOrColumnNotFoundError("sheet_not_found")

        worksheet = workbook[sheet_name]
        indexes: list[int] = []

        for row_number, row in enumerate(worksheet.iter_rows(values_only=True)):
            if row_number == 0:
                names = [_column_name(value, index) for index, value in enumerate(row)]
                try:
                    indexes = [names.index(column) for column in columns]
                except ValueError:
                    raise SheetOrColumnNotFoundError("column_not_found") from None
                continue

            # Profillemeyle aynı kural: tamamen boş satırlar Excel'in
            # "kullanılmış aralık" şişmesidir, satır sayılmaz.
            if all(
                value is None or (isinstance(value, str) and not value.strip()) for value in row
            ):
                continue

            yield tuple(_cell_text(row, index) for index in indexes)
    finally:
        workbook.close()


def _cell_text(row: tuple[Any, ...], index: int) -> str | None:
    if index >= len(row):
        return None
    cell = row[index]
    if cell is None:
        return None
    return cell if isinstance(cell, str) else str(cell)


def iter_column_values(
    path: Path,
    sheet_name: str,
    text_column: str,
) -> Iterator[str | None]:
    """Tek kolonluk kısayol; `iter_row_values` üzerinden aynı kuralları uygular."""
    for row in iter_row_values(path, sheet_name, [text_column]):
        yield row[0]


def validate_and_profile(path: Path, settings: Settings) -> dict[str, Any]:
    """Doğrulama + profilleme. Worker'ın çağırdığı tek giriş noktası."""
    uncompressed = validate_xlsx(path, settings)
    logger.info(
        "xlsx_validated",
        extra={"uncompressed_bytes": uncompressed},
    )
    return profile_xlsx(path, settings)
