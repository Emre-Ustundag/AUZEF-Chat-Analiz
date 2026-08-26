"""xlsx/JSON export testleri — plan §4 "Faz 4", ölçüt 2.

ALTYAPI GEREKTİRMEZ: üretilen çalışma kitabı openpyxl ile geri okunur.
Faz 4'ün asıl riski burada — "dosya indi" demek yetmez, dosyanın İÇİ
doğru olmalı.

Testlerin en önemlisi `test_sayilar_hucrede_sayi_olarak_durur`. Sayıyı
Türkçe biçimlendirilmiş bir dizeye çevirmek (`"12,5"`) Excel'de hücreyi
metne dönüştürür; toplama, sıralama, grafik ve pivot çalışmaz. Değer
karşılaştırması bunu YAKALAMAZ — `"12.5"` bir dize olarak da beklenen
sayıya "eşit görünebilir". Bu yüzden tip doğrudan sınanıyor.
"""

from __future__ import annotations

import json
import uuid
from datetime import UTC, datetime
from io import BytesIO

from openpyxl import load_workbook

from app.schemas.analysis import AnalysisMode, ConversationConfig, RowFilter
from app.schemas.report import (
    AnalysisReport,
    AnalysisWarning,
    PreprocessingSummary,
    SourceSummary,
    Theme,
    TokenUsage,
    TopQuestion,
)
from app.services import report_export


def build_report() -> AnalysisReport:
    """Gerçekçi ama küçük bir rapor: iki soru, iki tema, bir uyarı."""
    return AnalysisReport(
        analysis_id=uuid.UUID("11111111-2222-3333-4444-555555555555"),
        generated_at=datetime(2026, 8, 12, 10, 30, 0, tzinfo=UTC),
        source_summary=SourceSummary(
            filename="ornek.xlsx",
            sheet_name="Mesajlar",
            text_column="mesaj",
            total_rows=40,
        ),
        preprocessing_summary=PreprocessingSummary(
            analyzed_count=40,
            discarded_count=0,
            duplicate_count=35,
            redacted_count=2,
            unique_count=5,
        ),
        top_questions=[
            TopQuestion(
                id="q1",
                canonical_question="Sınav tarihleri ne zaman açıklanacak?",
                count=25,
                percentage=62.5,
                redacted_examples=[
                    "sınav tarihleri ne zaman açıklanacak",
                    "[EPOSTA] adresinden yazdım, sınav tarihi?",
                ],
            ),
            TopQuestion(
                id="q2",
                canonical_question="Harç ödemesi nasıl yapılır?",
                count=15,
                percentage=37.5,
                redacted_examples=["harç ödemesini nasıl yaparım"],
            ),
        ],
        themes=[
            Theme(id="t1", name="Sınav", count=25, percentage=62.5, related_question_ids=["q1"]),
            Theme(id="t2", name="Ödeme", count=15, percentage=37.5, related_question_ids=["q2"]),
        ],
        executive_summary="Kayıtların çoğunluğu sınav takvimiyle ilgili.",
        warnings=[
            AnalysisWarning(code="COST_LIMIT_APPROACHED", message="Maliyet sınırına yaklaşıldı.")
        ],
        model="anthropic/claude-sonnet-4.6",
        prompt_version="faq_analysis/v1",
        prompt_hash="sha256:abc123",
        token_usage=TokenUsage(prompt_tokens=1200, completion_tokens=300, total_tokens=1500),
        estimated_cost_usd=0.00825,
    )


def load(report: AnalysisReport | None = None) -> object:
    body = report_export.build_xlsx(report or build_report())
    return load_workbook(BytesIO(body))


# ------------------------------------------------------------------- yapı


def test_dosya_gercek_bir_xlsx() -> None:
    body = report_export.build_xlsx(build_report())
    # OOXML zip imzası: sahte/boş bir gövde buradan geçemez.
    assert body.startswith(b"PK\x03\x04")
    assert len(body) > 2_000


def test_sayfa_adlari_sabit() -> None:
    workbook = load()
    assert workbook.sheetnames == ["Özet", "Sorular", "Temalar"]  # type: ignore[attr-defined]


def test_sorular_sayfasinin_basligi_ve_satir_sayisi() -> None:
    report = build_report()
    sheet = load(report)["Sorular"]  # type: ignore[index]

    assert [cell.value for cell in sheet[1]] == [
        "Kimlik",
        "Soru",
        "Adet",
        "Oran (%)",
        "Örnek mesajlar (redakte)",
    ]
    # Başlık + her soru için bir satır. Fazlası kırpma/tekrar hatası demek.
    assert sheet.max_row == len(report.top_questions) + 1


def test_temalar_sayfasinin_basligi_ve_satir_sayisi() -> None:
    report = build_report()
    sheet = load(report)["Temalar"]  # type: ignore[index]

    assert [cell.value for cell in sheet[1]] == [
        "Kimlik",
        "Tema",
        "Adet",
        "Oran (%)",
        "İlgili soru kimlikleri",
    ]
    assert sheet.max_row == len(report.themes) + 1


# ------------------------------------------------- ASIL ÖLÇÜT: sayı tipi


def assert_numeric(value: object, label: str) -> None:
    """Hücrenin SAYI olduğunu doğrular.

    `float` beklemek YANLIŞ olurdu: xlsx sayıyı ondalık kısmı olmadan
    saklar ve openpyxl `62.5`'i `float`, `60.0`'ı `int` olarak geri okur.
    Kanıtlanmak istenen şey tip ayrıntısı değil, hücrenin metne
    dönüşmemiş olması — `bool` de dışlanıyor çünkü Python'da `int`
    sayılır ama Excel'de sayı değildir.
    """
    assert not isinstance(value, str), f"{label} metin olarak yazılmış: {value!r}"
    assert isinstance(value, int | float) and not isinstance(value, bool), (
        f"{label} sayı değil: {value!r} ({type(value).__name__})"
    )


def test_sayilar_hucrede_sayi_olarak_durur() -> None:
    """Adet/oran hücreleri METİN OLMAMALI (plan §4 Faz 4).

    Değer eşitliği tek başına yetmez: `"62,5"` veya `"62.5"` dizesi de
    gözle doğru görünür ama Excel onunla toplama, sıralama, grafik ve
    pivot yapamaz. Bu yüzden tip doğrudan sınanıyor.
    """
    report = build_report()
    workbook = load(report)

    questions = workbook["Sorular"]  # type: ignore[index]
    for row, question in zip(questions.iter_rows(min_row=2), report.top_questions, strict=True):
        assert_numeric(row[2].value, "adet")
        assert_numeric(row[3].value, "oran")
        assert row[2].value == question.count
        assert row[3].value == question.percentage

    themes = workbook["Temalar"]  # type: ignore[index]
    for row, theme in zip(themes.iter_rows(min_row=2), report.themes, strict=True):
        assert_numeric(row[2].value, "tema adedi")
        assert_numeric(row[3].value, "tema oranı")
        assert row[2].value == theme.count
        assert row[3].value == theme.percentage

    # Kesirli oran GERÇEKTEN kesirli kalmış olmalı: fixture 62.5/37.5
    # kullanıyor, yani tamsayıya yuvarlanma olsaydı burada patlardı.
    assert questions.cell(row=2, column=4).value == 62.5


def test_ozet_sayaclari_da_sayi() -> None:
    sheet = load()["Özet"]  # type: ignore[index]
    values = {row[0]: row[1] for row in sheet.iter_rows(min_row=2, values_only=True)}

    assert values["Toplam satır"] == 40
    assert_numeric(values["Toplam satır"], "toplam satır")
    assert_numeric(values["Analiz edilen kayıt"], "analiz edilen")
    assert_numeric(values["Toplam token"], "toplam token")
    assert_numeric(values["Maliyet (USD)"], "maliyet")
    assert_numeric(values["Cache'den okunan token"], "cache okuma token")
    assert_numeric(values["Cache'e yazılan token"], "cache yazma token")
    assert values["Maliyet (USD)"] == 0.00825


def test_ozet_satir_filtrelerini_yeniden_uretilebilir_bicimde_yazar() -> None:
    report = build_report()
    report = report.model_copy(
        update={
            "source_summary": report.source_summary.model_copy(
                update={
                    "row_filters": [
                        RowFilter(column="direction", allowed_values=["Kullanıcı"]),
                        RowFilter(column="message_type", allowed_values=["text", "free-text"]),
                    ]
                }
            )
        }
    )

    sheet = load(report)["Özet"]  # type: ignore[index]
    values = {row[0]: row[1] for row in sheet.iter_rows(min_row=2, values_only=True)}

    assert values["Satır filtreleri"] == ("direction = Kullanıcı; message_type = text | free-text")


def test_oranlar_rapordaki_degerlerle_tutar() -> None:
    """ADR §4: oranlar backend'de hesaplanır, export onları TÜRETMEZ."""
    report = build_report()
    sheet = load(report)["Sorular"]  # type: ignore[index]

    analyzed = report.preprocessing_summary.analyzed_count
    for row in sheet.iter_rows(min_row=2, values_only=True):
        count, percentage = row[2], row[3]
        assert percentage == round(count / analyzed * 100, 1)


# ------------------------------------------------------------- içerik/PII


def test_ornek_mesajlar_redakte_hâliyle_gider() -> None:
    """ADR §9: rapora giren örnekler zaten maskeli; export ham metne erişmez."""
    report = build_report()
    sheet = load(report)["Sorular"]  # type: ignore[index]

    examples = [row[4] for row in sheet.iter_rows(min_row=2, values_only=True)]
    assert all(text for text in examples), "örnek hücresi boş kalmamalı"

    birlesik = "\n".join(str(text) for text in examples)
    assert "[EPOSTA]" in birlesik
    # Rapor gövdesinde olmayan bir şey export'ta da olamaz.
    for question in report.top_questions:
        for example in question.redacted_examples:
            assert example in birlesik


def test_uyarilar_ozete_yazilir() -> None:
    sheet = load()["Özet"]  # type: ignore[index]
    metin = "\n".join(str(cell) for row in sheet.iter_rows(values_only=True) for cell in row)
    assert "COST_LIMIT_APPROACHED" in metin
    assert "Kayıtların çoğunluğu sınav takvimiyle ilgili." in metin


def test_contextual_metadata_xlsx_formulu_olamaz() -> None:
    """Kullanıcı denetimli eşleme değerleri Excel formülüne dönüşmemeli."""
    report = build_report()
    config = ConversationConfig(
        session_id_column="=1+1",
        message_order_column="+SUM(A1:A2)",
        role_column="@rol",
        message_type_column="-tur",
        user_role_values=["=Kullanıcı"],
        assistant_role_values=["+Bot"],
        target_message_types=["@text"],
        context_message_types=["@text", "-quick_reply"],
    )
    source = report.source_summary.model_copy(
        update={
            "analysis_mode": AnalysisMode.CONTEXTUAL_USER_TURNS,
            "conversation_config": config,
        }
    )
    report = report.model_copy(update={"source_summary": source})

    sheet = load(report)["Özet"]  # type: ignore[index]
    cells = {row[0].value: row[1] for row in sheet.iter_rows(min_row=2)}
    assert cells["Bot yanıtları bağlamda"].value == "Hayır"
    protected_labels = {
        "Session kolonu",
        "Mesaj sıra kolonu",
        "Rol kolonu",
        "Mesaj türü kolonu",
        "Kullanıcı rol değerleri",
        "Bot rol değerleri",
        "Hedef mesaj türleri",
        "Bağlam mesaj türleri",
    }

    for label in protected_labels:
        cell = cells[label]
        assert cell.data_type == "s"
        assert isinstance(cell.value, str)
        assert cell.value.startswith("'")


# ------------------------------------------------------------------ JSON


def test_json_govdesi_z_ile_biten_zaman_damgasi_tasir() -> None:
    """`/result` ile BİREBİR aynı serileştirme.

    Frontend `z.iso.datetime()` kullanıyor ve offset'li (`+00:00`) damgayı
    REDDEDİYOR. Export gövdesi `/result`'tan farklı serileşirse indirilen
    dosya sözleşme dışı kalır.
    """
    report = build_report()
    payload = json.loads(report.model_dump_json())

    assert payload["generated_at"].endswith("Z")
    assert payload["status"] == "completed"
    assert payload["top_questions"][0]["count"] == 25


# --------------------------------------------------------- dosya adı/başlık


def test_content_disposition_ascii_ve_kimlikten_turetilir() -> None:
    """Frontend yalnızca düz `filename="..."` biçimini ayrıştırıyor."""
    header = report_export.content_disposition("11111111-2222-3333-4444-555555555555", "xlsx")

    assert header == 'attachment; filename="analiz-11111111-2222-3333-4444-555555555555.xlsx"'
    header.encode("ascii")  # Türkçe karakter sızarsa burada patlar.


def test_bos_rapor_da_gecerli_dosya_uretir() -> None:
    """Soru/tema bulunamamış bir analiz export'u ÇÖKMEMELİ."""
    report = build_report().model_copy(update={"top_questions": [], "themes": [], "warnings": []})

    workbook = load(report)
    assert workbook.sheetnames == ["Özet", "Sorular", "Temalar"]  # type: ignore[attr-defined]
    assert workbook["Sorular"].max_row == 1  # type: ignore[index]
