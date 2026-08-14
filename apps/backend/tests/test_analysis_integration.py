"""Analiz akışının uçtan uca testi — plan §4 ölçütleri.

Postgres, Redis ve MinIO gerektirir (`docker compose up -d`).

`test_api_integration.py` ile aynı desen: Celery'ye gerçekten iş
gönderilmez, iş mantığı doğrudan `await` edilir. Broker üzerinden çalışmak
testi zamanlamaya bağımlı kılar ve rastgele patlar.
"""

from __future__ import annotations

import uuid
from collections.abc import AsyncIterator, Iterator
from json import dumps as json_dumps
from pathlib import Path
from typing import Any, cast

import httpx
import pytest
import pytest_asyncio
from celery.exceptions import SoftTimeLimitExceeded
from httpx import ASGITransport, AsyncClient
from sqlalchemy import text

from app.core.config import get_settings
from app.core.db import dispose_engine, session_scope
from app.main import create_app
from app.pipeline.llm_classifier import OpenRouterClassifier
from app.prompts.faq_analysis import get_prompt
from app.schemas.analysis import AnalysisJobRead
from app.schemas.report import AnalysisReport
from app.services import secret_store, storage
from app.services.openrouter import OpenRouterClient
from app.workers import tasks
from tests.fake_openrouter import FakeOpenRouter

pytestmark = pytest.mark.integration

FIXTURES = Path(__file__).parent / "fixtures"

TEST_KEY = "sk-or-v1-testkey0123456789abcdef0123456789"


@pytest.fixture(scope="module", autouse=True)
def _bucket() -> None:
    storage.ensure_bucket(get_settings())


@pytest.fixture(autouse=True)
def _no_broker(monkeypatch: pytest.MonkeyPatch) -> Iterator[None]:
    """Compose'daki worker testin kendi çağrısıyla yarışmasın."""
    monkeypatch.setattr(tasks.profile_upload, "delay", lambda *a, **k: None)
    monkeypatch.setattr(tasks.run_analysis_task, "delay", lambda *a, **k: None)
    yield


@pytest.fixture(autouse=True)
def provider(monkeypatch: pytest.MonkeyPatch) -> FakeOpenRouter:
    """`build_classifier`'ı sahte transport'a bağlar — GERÇEK AĞ YOK.

    Yamanın yeri bilinçli: `build_classifier` Faz 2'den beri TEK değişim
    noktası. Böylece test, `run_analysis`'in gerçek kodunu (anahtar okuma,
    maliyet tavanı, aşama ilerletme, hata çevirimi) baştan sona koşturur;
    yalnızca soketin öbür ucu sahte.
    """
    return install_fake_provider(monkeypatch)


def install_fake_provider(monkeypatch: pytest.MonkeyPatch, **kwargs: Any) -> FakeOpenRouter:
    """Sahte sağlayıcıyı kurar ve handler'ı döndürür.

    `provider` fixture'ı autouse olduğu için bu fonksiyon her testte EN AZ
    bir kez çalışır. Kendi senaryosunu kuran testler (bozuk JSON, rate
    limit ...) onu tekrar çağırır; aynı `monkeypatch` aynı özniteliği
    yeniden yamadığı için SON kurulum geçerli olur. Autouse olanı
    "gereksiz" diye kaldırmayın: senaryo kurmayan testlerin gerçek ağa
    çıkmasını o engelliyor.
    """
    fake = FakeOpenRouter(**kwargs)
    real_build = tasks.build_classifier

    def fake_build(**build_kwargs: Any) -> Any:
        settings = build_kwargs["settings"]
        client = OpenRouterClient(
            api_key=build_kwargs["api_key"],
            model=build_kwargs["model"],
            settings=settings,
            transport=httpx.MockTransport(fake),
            sleeper=lambda _s: None,
        )
        return OpenRouterClassifier(
            client=client,
            prompt=get_prompt(build_kwargs["prompt_version"]),
            model=build_kwargs["model"],
            settings=settings,
            on_progress=build_kwargs.get("on_progress"),
            # B5b: koşu sırasındaki tavan da AKTARILMALI. Aktarılmazsa
            # entegrasyon testleri korumayı hiç kurmaz ve "iş durdu mu?"
            # sorusunu sessizce yanlış cevaplar.
            max_cost_usd=build_kwargs.get("max_cost_usd"),
        )

    assert real_build is not fake_build
    monkeypatch.setattr(tasks, "build_classifier", fake_build)
    return fake


@pytest_asyncio.fixture
async def client() -> AsyncIterator[AsyncClient]:
    await dispose_engine()
    app = create_app()
    transport = ASGITransport(app=app)
    try:
        async with AsyncClient(transport=transport, base_url="http://test") as ac:
            yield ac
    finally:
        await dispose_engine()


async def _ready_upload(client: AsyncClient) -> tuple[uuid.UUID, dict[str, Any]]:
    """Geçerli bir dosya yükleyip profilini hazır hâle getirir."""
    path = FIXTURES / "valid_multi_sheet.xlsx"
    with path.open("rb") as handle:
        response = await client.post(
            "/api/v1/uploads",
            files={"file": (path.name, handle, "application/octet-stream")},
        )
    upload_id = uuid.UUID(response.json()["upload_id"])
    assert await tasks.run_upload_profiling(upload_id) == "ready"

    body = (await client.get(f"/api/v1/uploads/{upload_id}")).json()
    return upload_id, body["profile"]


def _request_body(upload_id: uuid.UUID, **overrides: Any) -> dict[str, Any]:
    payload: dict[str, Any] = {
        "upload_id": str(upload_id),
        "sheet_name": "Mesajlar",
        "text_column": "mesaj",
        "model": "anthropic/claude-sonnet-4.6",
        "prompt_version": "faq_analysis/v1",
        "top_n": 20,
        "max_cost_usd": 5.0,
    }
    payload.update(overrides)
    return payload


async def _create(
    client: AsyncClient,
    upload_id: uuid.UUID,
    *,
    key: str | None = TEST_KEY,
    **overrides: Any,
) -> Any:
    headers = {"X-OpenRouter-Key": key} if key is not None else {}
    return await client.post(
        "/api/v1/analyses",
        json=_request_body(upload_id, **overrides),
        headers=headers,
    )


# ------------------------------------------------------------- mutlu yol


async def test_analiz_uctan_uca_calisir(client: AsyncClient) -> None:
    upload_id, _ = await _ready_upload(client)

    created = await _create(client, upload_id)
    assert created.status_code == 202
    assert set(created.json()) == {"analysis_id", "status"}
    assert created.json()["status"] == "queued"

    analysis_id = uuid.UUID(created.json()["analysis_id"])

    # Başlangıç durumu sözleşmeye uymalı.
    job = (await client.get(f"/api/v1/analyses/{analysis_id}")).json()
    AnalysisJobRead.model_validate(job)
    assert job["status"] == "queued"
    assert job["progress"] == 0
    assert job["error"] is None
    assert job["created_at"].endswith("Z")

    assert await tasks.run_analysis(analysis_id) == "completed"

    job = (await client.get(f"/api/v1/analyses/{analysis_id}")).json()
    assert job["status"] == "completed"
    assert job["progress"] == 100
    assert job["estimated_seconds_remaining"] is None

    result = await client.get(f"/api/v1/analyses/{analysis_id}/result")
    assert result.status_code == 200

    report = result.json()
    parsed = AnalysisReport.model_validate(report)
    assert parsed.status == "completed"
    assert str(parsed.analysis_id) == str(analysis_id)
    assert parsed.source_summary.sheet_name == "Mesajlar"
    assert parsed.source_summary.text_column == "mesaj"
    # Fixture: 40 satır, 5 farklı mesaj tekrarlanıyor.
    assert parsed.source_summary.total_rows == 40
    assert parsed.preprocessing_summary.analyzed_count == 40
    assert parsed.preprocessing_summary.unique_count == 5

    # ADR §4: oranlar adetten türetilir.
    analyzed = parsed.preprocessing_summary.analyzed_count
    for question in parsed.top_questions:
        assert question.percentage == round(question.count / analyzed * 100, 1)

    # FAZ 3: token tüketimi artık GERÇEK — sağlayıcının `usage` bloğundan
    # geliyor, modelin metninden değil (ADR §4).
    assert parsed.token_usage.total_tokens > 0
    assert parsed.token_usage.total_tokens == (
        parsed.token_usage.prompt_tokens + parsed.token_usage.completion_tokens
    )
    assert parsed.estimated_cost_usd > 0.0
    # Model gerçekten çağrıldı ve prompt sürümü rapora işlendi.
    assert parsed.prompt_version == "faq_analysis/v1"
    assert parsed.prompt_hash.startswith("sha256:")

    await client.delete(f"/api/v1/uploads/{upload_id}")


async def test_rapor_pii_icermez(client: AsyncClient) -> None:
    """ADR §9: örnekler redakte edilmiş olarak rapora girer."""
    upload_id, _ = await _ready_upload(client)
    created = await _create(client, upload_id, sheet_name="Iletisim", text_column="not")
    analysis_id = uuid.UUID(created.json()["analysis_id"])

    assert await tasks.run_analysis(analysis_id) == "completed"
    report = (await client.get(f"/api/v1/analyses/{analysis_id}/result")).json()

    serialized = str(report)
    assert "ali@example.com" not in serialized
    assert "05551234567" not in serialized
    assert "12345678901" not in serialized
    assert "[EPOSTA]" in serialized
    assert report["preprocessing_summary"]["redacted_count"] > 0

    await client.delete(f"/api/v1/uploads/{upload_id}")


# ------------------------------------------------------------ hata yolları


async def test_anahtarsiz_istek_reddedilir(client: AsyncClient) -> None:
    upload_id, _ = await _ready_upload(client)
    response = await _create(client, upload_id, key=None)

    assert response.status_code == 422
    assert response.json()["code"] == "PROVIDER_AUTH_FAILED"

    await client.delete(f"/api/v1/uploads/{upload_id}")


async def test_whitelist_disi_model_reddedilir(client: AsyncClient) -> None:
    """ADR §9 değişmez 4: model yalnızca backend whitelist'inden seçilir."""
    upload_id, _ = await _ready_upload(client)
    response = await _create(client, upload_id, model="uydurma/model-9")

    assert response.status_code == 422
    assert response.json()["code"] == "SHEET_OR_COLUMN_NOT_FOUND"

    await client.delete(f"/api/v1/uploads/{upload_id}")


async def test_olmayan_kolon_reddedilir(client: AsyncClient) -> None:
    upload_id, _ = await _ready_upload(client)
    response = await _create(client, upload_id, text_column="olmayan_kolon")

    assert response.status_code == 422
    assert response.json()["code"] == "SHEET_OR_COLUMN_NOT_FOUND"

    await client.delete(f"/api/v1/uploads/{upload_id}")


async def test_olmayan_upload_404(client: AsyncClient) -> None:
    response = await _create(client, uuid.uuid4())
    assert response.status_code == 404
    assert response.json()["code"] == "JOB_NOT_FOUND"


async def test_bilinmeyen_analiz_404(client: AsyncClient) -> None:
    analysis_id = uuid.uuid4()
    assert (await client.get(f"/api/v1/analyses/{analysis_id}")).status_code == 404
    assert (await client.get(f"/api/v1/analyses/{analysis_id}/result")).status_code == 404
    assert (await client.delete(f"/api/v1/analyses/{analysis_id}")).status_code == 404


async def test_tamamlanmamis_isin_raporu_409(client: AsyncClient) -> None:
    """Devam eden iş için boş veri değil AÇIK hata dönülür (mock ile aynı)."""
    upload_id, _ = await _ready_upload(client)
    created = await _create(client, upload_id)
    analysis_id = uuid.UUID(created.json()["analysis_id"])

    response = await client.get(f"/api/v1/analyses/{analysis_id}/result")
    assert response.status_code == 409
    assert response.json()["code"] == "JOB_CONFLICT"

    await client.delete(f"/api/v1/uploads/{upload_id}")


# ------------------------------------------------------------------ iptal


async def test_iptal_edilen_is_cancelled_kalir(client: AsyncClient) -> None:
    """Plan §4 ölçüt 2 — worker iptali EZEMEZ."""
    upload_id, _ = await _ready_upload(client)
    created = await _create(client, upload_id)
    analysis_id = uuid.UUID(created.json()["analysis_id"])

    cancelled = await client.delete(f"/api/v1/analyses/{analysis_id}")
    assert cancelled.status_code == 204
    assert cancelled.content == b""

    job = (await client.get(f"/api/v1/analyses/{analysis_id}")).json()
    assert job["status"] == "cancelled"

    # Worker işi ŞİMDİ alsa bile durumu geri alamaz.
    assert await tasks.run_analysis(analysis_id) == "cancelled"

    job = (await client.get(f"/api/v1/analyses/{analysis_id}")).json()
    assert job["status"] == "cancelled"
    assert job["error"] is None
    # İptal edilen iş için rapor da olmamalı.
    assert (await client.get(f"/api/v1/analyses/{analysis_id}/result")).status_code == 409

    await client.delete(f"/api/v1/uploads/{upload_id}")


async def test_iptal_ilerlemeyi_sifirlamaz(client: AsyncClient) -> None:
    """B9 — iptal, işin nereye kadar geldiğini SİLMEMELİ.

    `cancel_analysis` durumu `cancelled` yaparken `progress`'i de
    `STAGE_PROGRESS[QUEUED]`'a (0.0) çekiyordu. Hata yolu (`_fail`) ise 100.0
    yazıyor — iki terminal yol birbiriyle tutarsızdı.

    Bugün arayüzde görünmüyor (`progress-screen.tsx`'in `cancelled` dalı
    ilerleme çubuğu render etmiyor), ama kayıt "iş %75'te iptal edildi"
    bilgisini taşımalı: iptali ne zaman aldığımızı DB'den okuyabilmek,
    ilerleme çubuğunu bir gün o ekrana koymaktan bağımsız olarak değerli.
    """
    upload_id, _ = await _ready_upload(client)
    created = await _create(client, upload_id)
    analysis_id = uuid.UUID(created.json()["analysis_id"])

    # İşi aşama ortasına taşı: iptal, buradaki değeri korumalı.
    async with session_scope() as session:
        await session.execute(
            text("UPDATE analyses SET status='analyzing', progress=75.0 WHERE id=:i"),
            {"i": str(analysis_id)},
        )
        await session.commit()

    assert (await client.delete(f"/api/v1/analyses/{analysis_id}")).status_code == 204

    job = (await client.get(f"/api/v1/analyses/{analysis_id}")).json()
    assert job["status"] == "cancelled"
    assert job["progress"] == 75.0, "iptal, işin geldiği noktayı silmemeli"

    await client.delete(f"/api/v1/uploads/{upload_id}")


async def test_biten_isin_iptali_409(client: AsyncClient) -> None:
    upload_id, _ = await _ready_upload(client)
    created = await _create(client, upload_id)
    analysis_id = uuid.UUID(created.json()["analysis_id"])

    assert await tasks.run_analysis(analysis_id) == "completed"

    response = await client.delete(f"/api/v1/analyses/{analysis_id}")
    assert response.status_code == 409
    assert response.json()["code"] == "JOB_CONFLICT"

    await client.delete(f"/api/v1/uploads/{upload_id}")


async def test_iki_kez_calisan_task_raporu_bozmaz(client: AsyncClient) -> None:
    """acks_late yeniden dağıtımı aynı işi iki kez gönderebilir."""
    upload_id, _ = await _ready_upload(client)
    created = await _create(client, upload_id)
    analysis_id = uuid.UUID(created.json()["analysis_id"])

    assert await tasks.run_analysis(analysis_id) == "completed"
    first = (await client.get(f"/api/v1/analyses/{analysis_id}/result")).json()

    assert await tasks.run_analysis(analysis_id) == "completed"
    second = (await client.get(f"/api/v1/analyses/{analysis_id}/result")).json()

    assert first == second

    await client.delete(f"/api/v1/uploads/{upload_id}")


# --------------------------------------------------------------- anahtar


async def test_anahtar_redise_sifreli_yazilir_ve_is_bitince_silinir(
    client: AsyncClient,
) -> None:
    """Plan §4 ölçüt 5 — PostgreSQL'de ve Redis'te düz anahtar YOK."""
    from redis import Redis

    settings = get_settings()
    upload_id, _ = await _ready_upload(client)
    created = await _create(client, upload_id)
    analysis_id = uuid.UUID(created.json()["analysis_id"])

    redis_client = Redis.from_url(settings.redis_url)
    try:
        stored = cast("bytes | None", redis_client.get(secret_store.redis_key(analysis_id)))
        assert stored is not None
        assert TEST_KEY.encode() not in stored
        ttl = cast("int", redis_client.ttl(secret_store.redis_key(analysis_id)))
        assert 0 < ttl <= settings.openrouter_key_ttl_seconds

        # PostgreSQL'de anahtarın hiçbir izi olmamalı.
        async with session_scope() as session:
            row = (
                await session.execute(
                    text("SELECT to_jsonb(a) AS data FROM analyses a WHERE a.id = :id"),
                    {"id": str(analysis_id)},
                )
            ).one()
        assert TEST_KEY not in str(row.data)
        assert "openrouter" not in str(row.data).lower()

        assert await tasks.run_analysis(analysis_id) == "completed"

        # ADR §9: iş bitince anahtar silinir.
        assert redis_client.get(secret_store.redis_key(analysis_id)) is None
    finally:
        redis_client.delete(secret_store.redis_key(analysis_id))
        redis_client.close()

    await client.delete(f"/api/v1/uploads/{upload_id}")


async def test_iptal_de_anahtari_siler(client: AsyncClient) -> None:
    from redis import Redis

    settings = get_settings()
    upload_id, _ = await _ready_upload(client)
    created = await _create(client, upload_id)
    analysis_id = uuid.UUID(created.json()["analysis_id"])

    redis_client = Redis.from_url(settings.redis_url)
    try:
        assert redis_client.get(secret_store.redis_key(analysis_id)) is not None
        await client.delete(f"/api/v1/analyses/{analysis_id}")
        assert redis_client.get(secret_store.redis_key(analysis_id)) is None
    finally:
        redis_client.delete(secret_store.redis_key(analysis_id))
        redis_client.close()

    await client.delete(f"/api/v1/uploads/{upload_id}")


# =====================================================================
# FAZ 3 — LLM yolları (hepsi sahte transport üzerinden, GERÇEK AĞ YOK)
# =====================================================================


async def test_maliyet_tavani_asiminda_is_failed_olur_ve_hic_cagri_yapilmaz(
    client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """ADR §9: "LLM çağrısı BAŞLAMADAN iş güvenli biçimde durur."

    Ölçüt iki katlı: iş `failed` olmalı VE sağlayıcıya tek bir istek bile
    gitmemeli. İkincisi olmadan "tavan" ancak faturayı gördükten sonra
    çalışan bir uyarı olurdu.
    """
    fake = install_fake_provider(monkeypatch)
    upload_id, _ = await _ready_upload(client)
    # Fixture'ın tahmini maliyeti bunun kat kat üzerinde.
    created = await _create(client, upload_id, max_cost_usd=0.0000001)
    analysis_id = uuid.UUID(created.json()["analysis_id"])

    assert await tasks.run_analysis(analysis_id) == "failed"

    # ASIL KANIT: hiçbir HTTP isteği yapılmadı.
    assert fake.requests == []

    job = (await client.get(f"/api/v1/analyses/{analysis_id}")).json()
    AnalysisJobRead.model_validate(job)
    assert job["status"] == "failed"
    assert job["error"]["code"] == "COST_LIMIT_EXCEEDED"
    assert job["error"]["status"] == 409
    # Kullanıcı ne yapacağını bilmeli: sayılar mesajda geçmeli.
    assert "USD" in job["error"]["detail"]
    # `retry_after` alanı HİÇ olmamalı (frontend .optional(), .nullable() değil).
    assert "retry_after" not in job["error"]

    await client.delete(f"/api/v1/uploads/{upload_id}")


async def test_iki_onarim_denemesinden_sonra_provider_bad_response(
    client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """ADR §8: geçersiz yanıt en fazla iki onarım denemesinden sonra sonlanır."""
    fake = install_fake_provider(monkeypatch, always_bad_json=True)
    upload_id, _ = await _ready_upload(client)
    created = await _create(client, upload_id)
    analysis_id = uuid.UUID(created.json()["analysis_id"])

    assert await tasks.run_analysis(analysis_id) == "failed"

    # 1 ilk deneme + 2 onarım = 3. Fazlası ADR §8'i çiğnerdi.
    assert len(fake.requests) == 3
    # Onarım turları modele KENDİ bozuk çıktısını geri vermiş olmalı.
    assert len(fake.requests[-1]["messages"]) == 6

    job = (await client.get(f"/api/v1/analyses/{analysis_id}")).json()
    assert job["status"] == "failed"
    assert job["error"]["code"] == "PROVIDER_BAD_RESPONSE"
    assert job["error"]["status"] == 502
    # ADR §9: ham model yanıtı hata cevabına SIZMAZ.
    assert "bu geçerli JSON değil" not in str(job["error"])

    await client.delete(f"/api/v1/uploads/{upload_id}")


async def test_gecici_bozuk_yanit_onarimla_duzelir(
    client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """İki onarım İÇİNDE düzelen model işi başarısız YAPMAMALI."""
    fake = install_fake_provider(monkeypatch, bad_json_first=2)
    upload_id, _ = await _ready_upload(client)
    created = await _create(client, upload_id)
    analysis_id = uuid.UUID(created.json()["analysis_id"])

    assert await tasks.run_analysis(analysis_id) == "completed"
    assert len(fake.requests) == 3

    report = (await client.get(f"/api/v1/analyses/{analysis_id}/result")).json()
    parsed = AnalysisReport.model_validate(report)
    # Onarım turlarının token'ı da faturaya girmeli.
    assert parsed.token_usage.total_tokens > 0

    await client.delete(f"/api/v1/uploads/{upload_id}")


async def test_rate_limit_tukendiginde_retry_after_ile_biter(
    client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """ADR §7: PROVIDER_RATE_LIMITED 429 + retry_after."""
    # openrouter_max_retries varsayılanı 4 → 5 istek; hepsi 429.
    fake = install_fake_provider(monkeypatch, rate_limit_first=99)
    upload_id, _ = await _ready_upload(client)
    created = await _create(client, upload_id)
    analysis_id = uuid.UUID(created.json()["analysis_id"])

    assert await tasks.run_analysis(analysis_id) == "failed"
    assert len(fake.requests) == get_settings().openrouter_max_retries + 1

    job = (await client.get(f"/api/v1/analyses/{analysis_id}")).json()
    AnalysisJobRead.model_validate(job)
    assert job["status"] == "failed"
    assert job["error"]["code"] == "PROVIDER_RATE_LIMITED"
    assert job["error"]["status"] == 429
    # Sağlayıcının söylediği süre kullanıcıya taşınmalı.
    assert job["error"]["retry_after"] == 3.0

    await client.delete(f"/api/v1/uploads/{upload_id}")


async def test_rate_limit_gecicilyse_is_tamamlanir(
    client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """Backoff işe yarıyor: geçici 429 işi öldürmemeli."""
    fake = install_fake_provider(monkeypatch, rate_limit_first=2)
    upload_id, _ = await _ready_upload(client)
    created = await _create(client, upload_id)
    analysis_id = uuid.UUID(created.json()["analysis_id"])

    assert await tasks.run_analysis(analysis_id) == "completed"
    assert len(fake.requests) == 3

    await client.delete(f"/api/v1/uploads/{upload_id}")


async def test_modelin_atladigi_kayitlar_rapordan_dusmez(
    client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """EN SİNSİ HATA — atlanan kayıt hiçbir toplama kontrolüne takılmaz."""
    install_fake_provider(monkeypatch, drop_records=2)
    upload_id, _ = await _ready_upload(client)
    created = await _create(client, upload_id)
    analysis_id = uuid.UUID(created.json()["analysis_id"])

    assert await tasks.run_analysis(analysis_id) == "completed"
    report = (await client.get(f"/api/v1/analyses/{analysis_id}/result")).json()
    parsed = AnalysisReport.model_validate(report)

    # Adetler yine analiz edilen kayda EŞİT olmalı.
    assert sum(q.count for q in parsed.top_questions) == (
        parsed.preprocessing_summary.analyzed_count
    )
    # Kullanıcı modelin kayıt atladığını GÖRMELİ.
    kodlar = {w.code for w in parsed.warnings}
    assert "LLM_UNASSIGNED_RECORDS" in kodlar

    await client.delete(f"/api/v1/uploads/{upload_id}")


async def test_anahtar_yalnizca_authorization_basliginda_gider(
    client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """ADR §9: anahtar prompt'a, gövdeye veya loga girmez."""
    fake = install_fake_provider(monkeypatch)
    upload_id, _ = await _ready_upload(client)
    created = await _create(client, upload_id)
    analysis_id = uuid.UUID(created.json()["analysis_id"])

    assert await tasks.run_analysis(analysis_id) == "completed"

    assert fake.auth_headers
    assert all(h == f"Bearer {TEST_KEY}" for h in fake.auth_headers)
    # İstek GÖVDESİNDE anahtar YOK.
    for body in fake.requests:
        assert TEST_KEY not in json_dumps(body)

    # Rapor gövdesinde de yok.
    report = (await client.get(f"/api/v1/analyses/{analysis_id}/result")).text
    assert TEST_KEY not in report

    await client.delete(f"/api/v1/uploads/{upload_id}")


async def test_anahtar_suresi_dolmussa_provider_auth_failed(
    client: AsyncClient,
) -> None:
    """Kuyrukta TTL'i dolmuş iş: traceback değil, sözleşmeye uygun hata."""
    from redis import Redis

    settings = get_settings()
    upload_id, _ = await _ready_upload(client)
    created = await _create(client, upload_id)
    analysis_id = uuid.UUID(created.json()["analysis_id"])

    # Anahtarı elle sil — TTL'in dolmasını taklit ediyor.
    redis_client = Redis.from_url(settings.redis_url)
    try:
        redis_client.delete(secret_store.redis_key(analysis_id))
    finally:
        redis_client.close()

    assert await tasks.run_analysis(analysis_id) == "failed"

    job = (await client.get(f"/api/v1/analyses/{analysis_id}")).json()
    assert job["error"]["code"] == "PROVIDER_AUTH_FAILED"
    assert job["error"]["status"] == 422

    await client.delete(f"/api/v1/uploads/{upload_id}")


async def test_prompt_kayitlari_delimiter_icinde_gonderir(
    client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """ADR §9: açık delimiter + tool çağrıları kapalı."""
    fake = install_fake_provider(monkeypatch)
    upload_id, _ = await _ready_upload(client)
    created = await _create(client, upload_id)
    analysis_id = uuid.UUID(created.json()["analysis_id"])

    assert await tasks.run_analysis(analysis_id) == "completed"

    body = fake.requests[0]
    assert "tools" not in body
    assert body["response_format"]["json_schema"]["strict"] is True
    prompt = body["messages"][1]["content"]
    assert "<kayitlar>" in prompt and "</kayitlar>" in prompt
    assert '<kayit id="' in prompt
    # System prompt backend'den geliyor, istemciden değil.
    assert "GÜVENLİK KURALLARI" in body["messages"][0]["content"]

    await client.delete(f"/api/v1/uploads/{upload_id}")


async def test_anahtar_loglara_yazilmaz(
    client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """Plan §4 ölçüt 5 / ADR §9: anahtar loglarda YOK.

    Tüm akış boyunca kök logger'a düşen HER kaydı yakalayıp anahtarı
    arıyoruz. `caplog` yerine elle handler takılıyor çünkü kanıtlanmak
    istenen şey tam olarak `RedactionFilter`'ın devrede olduğu yol.
    """
    import logging

    from app.core.logging import RedactionFilter

    captured: list[str] = []

    class _Capture(logging.Handler):
        """Filtre UYGULANDIKTAN SONRA kaydın tamamını yakalar.

        Yalnızca biçimlenmiş mesaja bakmak yetmez: anahtar `extra=` ile
        geçirilen bir alanda da taşınabilir. Bu yüzden `record.__dict__`'in
        tamamı da metne çevriliyor.
        """

        def emit(self, record: logging.LogRecord) -> None:
            captured.append(f"{record.getMessage()} {record.__dict__!r}")

    fake = install_fake_provider(monkeypatch)
    upload_id, _ = await _ready_upload(client)
    created = await _create(client, upload_id)
    analysis_id = uuid.UUID(created.json()["analysis_id"])

    handler = _Capture()
    handler.addFilter(RedactionFilter())
    root = logging.getLogger()
    root.addHandler(handler)
    previous = root.level
    root.setLevel(logging.DEBUG)
    try:
        assert await tasks.run_analysis(analysis_id) == "completed"
    finally:
        root.removeHandler(handler)
        root.setLevel(previous)

    logs = "\n".join(captured)
    assert logs, "log yakalanamadı; test bir şey kanıtlamıyor"
    assert TEST_KEY not in logs
    assert "sk-or-v1" not in logs
    # Sağlayıcıya giden istek gerçekten yapıldı — yani test boş bir akışı
    # değil, anahtarın kullanıldığı akışı ölçüyor.
    assert fake.auth_headers

    await client.delete(f"/api/v1/uploads/{upload_id}")


# =====================================================================
# FAZ 4 — export
# =====================================================================


async def test_export_json_result_govdesiyle_ayni(client: AsyncClient) -> None:
    """Plan §4 Faz 4: `format=json` rapor gövdesini attachment olarak verir."""
    upload_id, _ = await _ready_upload(client)
    created = await _create(client, upload_id)
    analysis_id = uuid.UUID(created.json()["analysis_id"])
    assert await tasks.run_analysis(analysis_id) == "completed"

    response = await client.get(f"/api/v1/analyses/{analysis_id}/export?format=json")
    assert response.status_code == 200
    assert response.headers["content-type"].startswith("application/json")
    assert response.headers["content-disposition"] == (
        f'attachment; filename="analiz-{analysis_id}.json"'
    )

    # İndirilen gövde `/result` ile BİREBİR aynı olmalı: ayrışırsa kullanıcı
    # ekranda gördüğünden farklı bir dosya indirmiş olur.
    result = (await client.get(f"/api/v1/analyses/{analysis_id}/result")).json()
    assert response.json() == result
    # Frontend `z.iso.datetime()` offset'li damgayı reddeder.
    assert response.json()["generated_at"].endswith("Z")

    await client.delete(f"/api/v1/uploads/{upload_id}")


async def test_export_xlsx_gercek_dosya_ve_sayilar_raporla_tutar(client: AsyncClient) -> None:
    """Plan §4 Faz 4 ölçüt 2 — indirilen dosya openpyxl ile açılıp doğrulanır."""
    from io import BytesIO

    from openpyxl import load_workbook

    upload_id, _ = await _ready_upload(client)
    created = await _create(client, upload_id)
    analysis_id = uuid.UUID(created.json()["analysis_id"])
    assert await tasks.run_analysis(analysis_id) == "completed"

    response = await client.get(f"/api/v1/analyses/{analysis_id}/export?format=xlsx")
    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert response.headers["content-disposition"] == (
        f'attachment; filename="analiz-{analysis_id}.xlsx"'
    )
    assert response.content.startswith(b"PK\x03\x04")

    report = AnalysisReport.model_validate(
        (await client.get(f"/api/v1/analyses/{analysis_id}/result")).json()
    )

    workbook = load_workbook(BytesIO(response.content))
    assert workbook.sheetnames == ["Özet", "Sorular", "Temalar"]

    questions = workbook["Sorular"]
    assert questions.max_row == len(report.top_questions) + 1
    for row, question in zip(questions.iter_rows(min_row=2), report.top_questions, strict=True):
        assert row[0].value == question.id
        assert row[2].value == question.count
        assert row[3].value == question.percentage
        # Sayı METİN olmamalı — yoksa Excel'de toplama/grafik çalışmaz.
        assert not isinstance(row[2].value, str)
        assert not isinstance(row[3].value, str)

    themes = workbook["Temalar"]
    assert themes.max_row == len(report.themes) + 1
    for row, theme in zip(themes.iter_rows(min_row=2), report.themes, strict=True):
        assert row[2].value == theme.count
        assert row[3].value == theme.percentage

    await client.delete(f"/api/v1/uploads/{upload_id}")


async def test_export_ornekleri_redakte_gider(client: AsyncClient) -> None:
    """ADR §9: xlsx'e ham öğrenci mesajı değil, maskelenmiş örnek girer."""
    from io import BytesIO

    from openpyxl import load_workbook

    upload_id, _ = await _ready_upload(client)
    created = await _create(client, upload_id, sheet_name="Iletisim", text_column="not")
    analysis_id = uuid.UUID(created.json()["analysis_id"])
    assert await tasks.run_analysis(analysis_id) == "completed"

    response = await client.get(f"/api/v1/analyses/{analysis_id}/export?format=xlsx")
    workbook = load_workbook(BytesIO(response.content))
    metin = "\n".join(
        str(cell)
        for sheet in workbook.worksheets
        for row in sheet.iter_rows(values_only=True)
        for cell in row
        if cell is not None
    )

    assert "ali@example.com" not in metin
    assert "05551234567" not in metin
    assert "12345678901" not in metin
    assert "[EPOSTA]" in metin

    await client.delete(f"/api/v1/uploads/{upload_id}")


async def test_tamamlanmamis_analizde_export_409(client: AsyncClient) -> None:
    """Mock ile aynı davranış: rapor hazır değilse JOB_CONFLICT."""
    upload_id, _ = await _ready_upload(client)
    created = await _create(client, upload_id)
    analysis_id = uuid.UUID(created.json()["analysis_id"])

    for fmt in ("json", "xlsx"):
        response = await client.get(f"/api/v1/analyses/{analysis_id}/export?format={fmt}")
        assert response.status_code == 409, fmt
        assert response.json()["code"] == "JOB_CONFLICT"

    await client.delete(f"/api/v1/uploads/{upload_id}")


async def test_bilinmeyen_analizde_export_404(client: AsyncClient) -> None:
    response = await client.get(f"/api/v1/analyses/{uuid.uuid4()}/export?format=xlsx")
    assert response.status_code == 404
    assert response.json()["code"] == "JOB_NOT_FOUND"


async def test_gecersiz_format_json_dondurur(client: AsyncClient) -> None:
    """Mock `exportFormatSchema.catch("json")` kullanıyor; backend de öyle.

    `Literal[...]` doğrulaması burada 415 `UPLOAD_INVALID_TYPE` üretirdi ve
    kullanıcı yüklediği dosyayla ilgisi olmayan "Desteklenmeyen dosya türü"
    mesajını görürdü.
    """
    upload_id, _ = await _ready_upload(client)
    created = await _create(client, upload_id)
    analysis_id = uuid.UUID(created.json()["analysis_id"])
    assert await tasks.run_analysis(analysis_id) == "completed"

    for query in ("?format=pdf", "?format=", ""):
        response = await client.get(f"/api/v1/analyses/{analysis_id}/export{query}")
        assert response.status_code == 200, query
        assert response.headers["content-type"].startswith("application/json"), query

    await client.delete(f"/api/v1/uploads/{upload_id}")


async def test_bilinmeyen_prompt_surumu_reddedilir(
    client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """Prompt metni backend'de sürümlenir; sessizce varsayılana düşülmez."""
    install_fake_provider(monkeypatch)
    upload_id, _ = await _ready_upload(client)
    created = await _create(client, upload_id, prompt_version="faq_analysis/v99")
    analysis_id = uuid.UUID(created.json()["analysis_id"])

    assert await tasks.run_analysis(analysis_id) == "failed"
    job = (await client.get(f"/api/v1/analyses/{analysis_id}")).json()
    assert job["error"]["code"] == "SHEET_OR_COLUMN_NOT_FOUND"

    await client.delete(f"/api/v1/uploads/{upload_id}")


# ------------------------------------------------- B2: soft timeout (ADR §2)


async def test_soft_timeout_siniflandirmada_zaman_asimi_kodu_verir(
    client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """B2 — 45 dakikalık hard timeout kendi kodunu almalı.

    `config.py`'deki yorum soft limitin amacını "işi sözleşmeye uygun bir hata
    ile kapatmak" diye tanımlıyor ama `SoftTimeLimitExceeded` için hiçbir
    handler yoktu. `Exception` alt sınıfı olduğu için sınıflandırmayı saran
    geniş handler onu yakalıyor ve `PROVIDER_BAD_RESPONSE` yazıyordu:
    kullanıcı 45 dakika bekledikten sonra "Dil modelinden geçerli bir yanıt
    alınamadı" görüyordu. Doğru kod `PROVIDER_TIMEOUT` ve Türkçe metni tam
    olarak bunu anlatıyor ("Dil modeli zaman aşımına uğradı").

    Ölçülen şey yalnızca kod değil: işin TERMİNAL kapandığı da doğrulanıyor.
    """
    install_fake_provider(monkeypatch)
    upload_id, _ = await _ready_upload(client)
    created = await _create(client, upload_id)
    analysis_id = uuid.UUID(created.json()["analysis_id"])

    real_build = tasks.build_classifier

    def build_with_timeout(**kwargs: Any) -> Any:
        classifier = real_build(**kwargs)

        def raise_soft_timeout(*args: Any, **kw: Any) -> Any:
            raise SoftTimeLimitExceeded

        monkeypatch.setattr(classifier, "classify", raise_soft_timeout, raising=False)
        return classifier

    monkeypatch.setattr(tasks, "build_classifier", build_with_timeout)

    assert await tasks.run_analysis(analysis_id) == "failed"

    job = (await client.get(f"/api/v1/analyses/{analysis_id}")).json()
    assert job["status"] == "failed"
    assert job["error"]["code"] == "PROVIDER_TIMEOUT"
    assert job["error"]["status"] == 504

    await client.delete(f"/api/v1/uploads/{upload_id}")


async def test_soft_timeout_on_islemede_de_zaman_asimi_kodu_verir(
    client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """Aynı ayrım ön işlemede de geçerli olmalı.

    100.000 satırlık bir dosyayı okumak da limite yaklaşabilir; oradaki geniş
    handler `INTERNAL_ERROR` ("Beklenmeyen bir hata oluştu") yazıyordu, ki
    yapılandırılmış bir zaman aşımı beklenmeyen bir şey değil.
    """
    install_fake_provider(monkeypatch)
    upload_id, _ = await _ready_upload(client)
    created = await _create(client, upload_id)
    analysis_id = uuid.UUID(created.json()["analysis_id"])

    def raise_soft_timeout(*args: Any, **kwargs: Any) -> Any:
        raise SoftTimeLimitExceeded

    monkeypatch.setattr(tasks, "iter_column_values", raise_soft_timeout)

    assert await tasks.run_analysis(analysis_id) == "failed"

    job = (await client.get(f"/api/v1/analyses/{analysis_id}")).json()
    assert job["error"]["code"] == "PROVIDER_TIMEOUT"

    await client.delete(f"/api/v1/uploads/{upload_id}")


async def test_kosu_ortasinda_tavan_asilirsa_is_durur(
    client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """B5b — worker seviyesinde: koşu ortasında kesilen iş kendi kodunu alır.

    Sınıflandırıcı testi korumanın çalıştığını gösteriyor; bu test BAĞLANTININ
    kurulduğunu gösteriyor: tavan `build_classifier`'a geçiyor mu, istisna
    `_fail`'e çevriliyor mu, kod sözleşmedeki 13. kod mu.

    `COST_LIMIT_EXCEEDED`: sözleşmedeki tek maliyet tavanı kodu hem ön kontrolü
    harcanmadı" demek. Burada para harcandı ve kullanıcının bunu bilmesi
    gerekiyor.
    """
    # Sağlayıcı her çağrıda pahalı: ilk chunk'tan sonra tavan aşılır.
    install_fake_provider(monkeypatch, prompt_tokens=400_000, completion_tokens=0)
    upload_id, _ = await _ready_upload(client)
    # sonnet-5 girdi 2.0 USD/M -> çağrı başına 0.8 USD; tavan 0.5 USD.
    created = await _create(client, upload_id, max_cost_usd=0.5)
    analysis_id = uuid.UUID(created.json()["analysis_id"])

    assert await tasks.run_analysis(analysis_id) == "failed"

    job = (await client.get(f"/api/v1/analyses/{analysis_id}")).json()
    assert job["error"]["code"] == "COST_LIMIT_EXCEEDED"
    assert job["error"]["status"] == 409

    await client.delete(f"/api/v1/uploads/{upload_id}")
