"""Test fixture üreticileri.

İki tür fixture var:

* `tests/fixtures/` altında COMMIT EDİLENLER — küçük ve deterministik
  dosyalar. `python -m tests.factories` ile yeniden üretilirler.
* Çalışma anında üretilenler — ZIP bombası gibi commit edilmesi anlamsız
  veya sakıncalı olanlar. Bunlar `tmp_path` içine yazılır.

Şifreli dosya fixture'ı hakkında DÜRÜST NOT: gerçek bir parola korumalı
`.xlsx` üretmek ECMA-376 agile encryption uygulamayı gerektirir ve bunu
yapan yazan bir Python kütüphanesi yok (msoffcrypto-tool yalnızca okur).
Bu yüzden fixture, parola korumalı bir OOXML dosyasının AYIRT EDİCİ
ÖZELLİĞİNİ taşır: dosya `PK` ile değil, OLE2/CFB imzasıyla başlar. Doğrulama
kodunun şifreli dosyayı yakaladığı nokta zaten tam olarak burasıdır, çünkü
şifreli paket OOXML zip'ini bir OLE2 kabına sarar.
"""

from __future__ import annotations

import io
import zipfile
from pathlib import Path

from openpyxl import Workbook

FIXTURES_DIR = Path(__file__).parent / "fixtures"

#: OLE2/CFB kap imzası — parola korumalı xlsx ve eski .xls bununla başlar.
OLE2_HEADER = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"

_CONTENT_TYPES = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="xml" ContentType="application/xml"/>
  <Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>
</Types>"""

_WORKBOOK_XML = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <sheets><sheet name="Sayfa1" sheetId="1" r:id="rId1"
    xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"/></sheets>
</workbook>"""


def build_valid_workbook(path: Path) -> Path:
    """Çok sayfalı, gerçekçi kolon karışımı olan geçerli bir .xlsx üretir.

    Kolonlar bilinçli olarak farklı tiplerde: `is_likely_text` sezgiselinin
    metin kolonunu tarih/sayı/kategori kolonlarından ayırabildiğini test
    edebilmek için.
    """
    from datetime import datetime

    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Mesajlar"
    sheet.append(["tarih", "kullanici_id", "mesaj", "kanal"])

    messages = [
        "sınav tarihleri ne zaman açıklanacak acaba bilgi alabilir miyim",
        "ders materyallerine nereden ulaşabilirim linkler çalışmıyor",
        "harç ödemesini nasıl yaparım banka şubesinden mi",
        "kayıt yenileme işlemi için son tarih ne zaman",
        "sınav yerimi nereden öğrenebilirim bilgi yok",
    ]
    for index in range(40):
        sheet.append(
            [
                datetime(2026, 3, 1, 9, (index % 50)),
                1000 + index,
                messages[index % len(messages)],
                "web" if index % 2 == 0 else "mobil",
            ]
        )

    second = workbook.create_sheet("Ham Veri")
    second.append(["id", "icerik"])
    for index in range(10):
        second.append([index + 1, f"kayıt {index + 1} serbest metin içeriği burada"])

    # PII redaksiyonunun profil örneklerinde gerçekten çalıştığını
    # doğrulayabilmek için bilinçli olarak PII içeren bir sayfa.
    third = workbook.create_sheet("Iletisim")
    third.append(["not"])
    third.append(["ogrenci ali@example.com adresinden yazdi ve 05551234567 numarasini verdi"])
    third.append(["tc kimlik 12345678901 ile kayit yaptirdi bilgi guncellensin lutfen"])

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path


def _minimal_ooxml_members() -> dict[str, str]:
    return {
        "[Content_Types].xml": _CONTENT_TYPES,
        "xl/workbook.xml": _WORKBOOK_XML,
    }


def build_macro_enabled(path: Path) -> Path:
    """Yapısal olarak GERÇEK bir makrolu çalışma kitabı.

    `.xlsx` uzantısı taşır ama içinde `xl/vbaProject.bin` vardır — ayırt edici
    işaret uzantı değil, tam olarak bu üyedir.
    """
    path.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as archive:
        for name, content in _minimal_ooxml_members().items():
            archive.writestr(name, content)
        archive.writestr("xl/vbaProject.bin", b"\x00\x01\x02CMG=\"macro payload\"")
    return path


def build_corrupt_zip(path: Path) -> Path:
    """`PK\\x03\\x04` ile başlayan ama içeriği bozuk bir dosya.

    Magic bytes kontrolünü GEÇER; asıl reddin zip dizini okunurken gelmesi
    gerekir. İmza kontrolüne güvenip zip ayrıştırmasını atlayan bir uygulama
    bu dosyada çöker.
    """
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(b"PK\x03\x04" + b"\xde\xad\xbe\xef" * 64)
    return path


def build_fake_encrypted(path: Path) -> Path:
    """Parola korumalı xlsx'in imzasını taşıyan dosya (bkz. modül başlığı)."""
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(OLE2_HEADER + b"\x00" * 504)
    return path


def build_empty_file(path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(b"")
    return path


def build_zip_bomb(path: Path, *, uncompressed_bytes: int = 512 * 1024 * 1024) -> Path:
    """Küçük ama açıldığında yüz megabaytlara ulaşan bir arşiv.

    Commit EDİLMEZ, test sırasında üretilir. Sıfır baytlar deflate ile ~1000:1
    oranında sıkıştığı için dosya birkaç yüz KB kalır.
    """
    path.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED, compresslevel=9) as archive:
        for name, content in _minimal_ooxml_members().items():
            archive.writestr(name, content)
        archive.writestr("xl/worksheets/sheet1.xml", b"\x00" * uncompressed_bytes)
    return path


def build_lying_zip(path: Path) -> Path:
    """Zip dizininde KÜÇÜK boyut beyan eden, gerçekte çok daha büyük üye.

    Beyan edilen boyutlara güvenen bir bomba kontrolünü atlatır; yalnızca
    açılan baytları sayan ikinci katman yakalayabilir.
    """
    path.parent.mkdir(parents=True, exist_ok=True)
    payload = b"\x00" * (64 * 1024 * 1024)

    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED, compresslevel=9) as archive:
        for name, content in _minimal_ooxml_members().items():
            archive.writestr(name, content)
        archive.writestr("xl/worksheets/sheet1.xml", payload)

    raw = bytearray(buffer.getvalue())
    # Merkezi dizindeki ve yerel başlıktaki `uncompressed size` alanlarını
    # yalana çeviriyoruz: 64 MB yerine 1 KB yazıyoruz.
    real_size = len(payload).to_bytes(4, "little")
    fake_size = (1024).to_bytes(4, "little")
    raw = bytearray(raw.replace(real_size, fake_size))

    path.write_bytes(bytes(raw))
    return path


def main() -> None:
    build_valid_workbook(FIXTURES_DIR / "valid_multi_sheet.xlsx")
    build_macro_enabled(FIXTURES_DIR / "macro_enabled.xlsx")
    build_corrupt_zip(FIXTURES_DIR / "corrupt.xlsx")
    build_fake_encrypted(FIXTURES_DIR / "encrypted.xlsx")
    build_empty_file(FIXTURES_DIR / "empty.xlsx")
    print(f"Fixture'lar yazıldı: {FIXTURES_DIR}")


if __name__ == "__main__":
    main()
