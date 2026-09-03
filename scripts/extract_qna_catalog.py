"""Eski botun kendi soru katalogunu QuickReplyLabel'dan cikarir.

"QnA secimi: ..." etiketleri, botun kullaniciya sundugu hazir sorulardir.
Insan eliyle yazilmislar ve tiklama sayilari gercek talebi olcer — LLM'e
uretilecek bir taksonomiden daha saglam bir cekirdek.

Cikti: outputs/analiz/qna-katalog-yil.json  (soru, tiklama, aylik dagilim)
"""
from __future__ import annotations

import collections
import json
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "chatbot.xlsx"
DST = ROOT / "outputs" / "analiz" / "qna-katalog-yil.json"
PREFIX = "QnA seçimi:"


def main() -> int:
    wb = load_workbook(SRC, read_only=True)
    clicks: collections.Counter[str] = collections.Counter()
    per_month: dict[str, collections.Counter[str]] = collections.defaultdict(collections.Counter)
    total = 0
    for name in wb.sheetnames:
        it = wb[name].iter_rows(values_only=True)
        idx = {h: i for i, h in enumerate(next(it))}
        for row in it:
            if row is None:
                continue
            total += 1
            q = str(row[idx["QuickReplyLabel"]] or "").strip()
            if not q.startswith(PREFIX):
                continue
            question = q[len(PREFIX):].strip()
            if not question:
                continue
            clicks[question] += 1
            per_month[question][str(row[idx["SessionStartTr"]] or "")[:7]] += 1
            if total % 1_000_000 == 0:
                print(f"  ...{total} satır, {len(clicks)} soru", flush=True)

    out = [
        {
            "question": q,
            "clicks": c,
            "months": dict(sorted(per_month[q].items())),
        }
        for q, c in clicks.most_common()
    ]
    DST.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")

    tot = sum(clicks.values())
    print(f"\n{'=' * 66}")
    print(f"QnA katalogu: {len(clicks)} soru, {tot} tiklama")
    cum = 0
    n80 = n90 = 0
    for i, (_, c) in enumerate(clicks.most_common(), 1):
        cum += c
        if not n80 and cum / tot >= 0.80:
            n80 = i
        if not n90 and cum / tot >= 0.90:
            n90 = i
    print(f"  tiklamalarin %80'i ilk {n80} soruda")
    print(f"  tiklamalarin %90'i ilk {n90} soruda")
    print(f"  5'ten az tiklanan: {sum(1 for c in clicks.values() if c < 5)}")
    print(f"\nEN COK TIKLANAN 20:")
    for q, c in clicks.most_common(20):
        print(f"  {c:>6}  {q[:80]}")
    print(f"\nkaydedildi: {DST}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
