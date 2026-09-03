"""FAZ 0 — chatbot.xlsx'i session düzeyinde çalışma dosyasına indirger.

Deterministik, LLM yok. 7 sayfa tek akışta okunur, belleğe alınmaz.
Her yazılı session için: birleşik kullanıcı metni, ay, kanal ve sonuç
sinyalleri (onay/ret/anlayamadım sayıları).

Çıktı: outputs/analiz/sessions-yil.jsonl
"""
from __future__ import annotations

import collections
import json
import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "apps" / "backend"))
from app.pipeline.preprocess import COURTESY_ONLY, is_system_message, normalize  # noqa: E402
from app.services.redaction import redact_pii  # noqa: E402
from openpyxl import load_workbook  # noqa: E402

SRC = ROOT / "chatbot.xlsx"
DST = ROOT / "outputs" / "analiz" / "sessions-yil.jsonl"
WS = re.compile(r"\s+")
FALLBACK = "sizi ne yazık ki anlayamadım"
MIN_LEN = 3


class S:
    __slots__ = ("turns", "ok", "rej", "fb", "month", "channel")

    def __init__(self, month: str, channel: str) -> None:
        self.turns: list[str] = []
        self.ok = 0
        self.rej = 0
        self.fb = 0
        self.month = month
        self.channel = channel


def main() -> int:
    wb = load_workbook(SRC, read_only=True)
    sess: dict[str, S] = {}
    total = 0
    for name in wb.sheetnames:
        it = wb[name].iter_rows(values_only=True)
        idx = {h: i for i, h in enumerate(next(it))}
        for row in it:
            if row is None:
                continue
            total += 1
            sid = row[idx["SessionId"]]
            if not sid:
                continue
            sid = str(sid)
            st = sess.get(sid)
            if st is None:
                st = sess[sid] = S(
                    str(row[idx["SessionStartTr"]] or "")[:7],
                    str(row[idx["Channel"]] or ""),
                )
            text = WS.sub(" ", str(row[idx["MessageTextClean"]] or "")).strip()
            direction = str(row[idx["Direction"]] or "")
            if direction == "Bot":
                if text.casefold().startswith(FALLBACK):
                    st.fb += 1
                continue
            q = str(row[idx["QuickReplyLabel"]] or "").strip()
            if q == "Onayladı":
                st.ok += 1
                continue
            if q == "Reddetti":
                st.rej += 1
                continue
            if str(row[idx["MessageType"]] or "") != "text":
                continue
            if (
                not text
                or text == "None"
                or len(text) < MIN_LEN
                or is_system_message(text)
                or normalize(text) in COURTESY_ONLY
            ):
                continue
            st.turns.append(redact_pii(text))
            if total % 1_000_000 == 0:
                print(f"  ...{total} satır, {len(sess)} session", flush=True)

    DST.parent.mkdir(parents=True, exist_ok=True)
    months: collections.Counter[str] = collections.Counter()
    written = 0
    with DST.open("w", encoding="utf-8") as fh:
        for sid, st in sess.items():
            if not st.turns:
                continue
            written += 1
            months[st.month] += 1
            fh.write(json.dumps({
                "session": sid, "month": st.month, "channel": st.channel,
                "text": " / ".join(st.turns), "turns": len(st.turns),
                "ok": st.ok, "rej": st.rej, "fallback": st.fb,
            }, ensure_ascii=False) + "\n")

    print(f"\n{'=' * 66}")
    print(f"satır {total} | session {len(sess)} | YAZILI SESSION {written} "
          f"(%{written / len(sess) * 100:.1f})")
    print(f"çıktı: {DST}  ({DST.stat().st_size / 1024 / 1024:.1f} MB)")
    print("\nAYLIK YAZILI SESSION:")
    for m, c in sorted(months.items()):
        print(f"   {m}: {c:>7}  {'#' * int(c / max(months.values()) * 40)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
