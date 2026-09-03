"""chatbot.xlsx (7 sayfa, ~7M satır) tam profili — akışlı, LLM yok."""
from __future__ import annotations

import collections
import sys
from pathlib import Path

from openpyxl import load_workbook

SRC = Path(__file__).resolve().parents[1] / "chatbot.xlsx"


def main() -> int:
    wb = load_workbook(SRC, read_only=True)
    months: collections.Counter[str] = collections.Counter()
    per_sheet: dict[str, int] = {}
    direction: collections.Counter[str] = collections.Counter()
    mtype: collections.Counter[str] = collections.Counter()
    qlabel: collections.Counter[str] = collections.Counter()
    feedback: collections.Counter[str] = collections.Counter()
    sessions: set[str] = set()
    lo = hi = None
    total = 0

    for name in wb.sheetnames:
        ws = wb[name]
        it = ws.iter_rows(values_only=True)
        hdr = list(next(it))
        idx = {h: i for i, h in enumerate(hdr)}
        n = 0
        for row in it:
            if row is None or len(row) <= idx.get("SessionId", 0):
                continue
            n += 1
            total += 1
            sid = row[idx["SessionId"]]
            if sid:
                sessions.add(str(sid))
            d = str(row[idx["SessionStartTr"]] or "")[:10]
            if d and d[0].isdigit():
                months[d[:7]] += 1
                if lo is None or d < lo:
                    lo = d
                if hi is None or d > hi:
                    hi = d
            direction[str(row[idx["Direction"]])] += 1
            mtype[str(row[idx["MessageType"]])] += 1
            q = str(row[idx["QuickReplyLabel"]] or "").strip()
            if q and q != "None":
                qlabel[q if q in ("Onayladı", "Reddetti") else
                       ("QnA seçimi" if q.startswith("QnA seçimi") else "(diğer etiket)")] += 1
            f = str(row[idx["MessageFeedback"]] or "").strip()
            if f and f != "None":
                feedback[f] += 1
            if total % 500000 == 0:
                print(f"  ...{total} satır", flush=True)
        per_sheet[name] = n
        print(f"sayfa {name}: {n} satır", flush=True)

    print(f"\n{'=' * 70}")
    print(f"TOPLAM SATIR : {total}")
    print(f"SESSION      : {len(sessions)}")
    print(f"TARİH ARALIĞI: {lo} → {hi}")
    print(f"\nAY DAĞILIMI:")
    for m, c in sorted(months.items()):
        print(f"   {m}: {c:>9}  {'#' * int(c / max(months.values()) * 40)}")
    print(f"\ndirection: {dict(direction)}")
    print(f"message_type: {dict(mtype.most_common(10))}")
    print(f"\nQuickReplyLabel: {dict(qlabel)}")
    print(f"MessageFeedback: {dict(feedback.most_common(6)) or 'BOŞ'}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
