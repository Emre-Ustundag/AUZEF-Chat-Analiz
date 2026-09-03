"""KB'de karsiligi olmayan sorular icin ice aktarilabilir taslak uretir.

Cikti module_479_Qna.csv ile AYNI formatta: question; answer; tags;
query_1..20. Fark: `answer` BOS birakilir — AUZEF'e ozgu gercekleri
uydurmak, bos birakmaktan cok daha kotudur.

Doldurulanlar:
  question   bosluk maddesinin kanonik hali
  query_1..20 O maddeye ATANMIS GERCEK ogrenci cumleleri (uydurma degil)
  tags       mevcut KB'nin tag sozlugunden en yakin kaydinki
  ek kolonlar (import oncesi silinecek): hacim, sorunlu oran ve eski botun
             o session'larda verdigi en sik anlamli cevap — taslak icin
"""

from __future__ import annotations

import collections
import csv
import json
import os
import re
import sys
import unicodedata
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))
from profile_new_kb import load as load_kb  # noqa: E402

GAP = ROOT / "outputs" / "analiz" / "kb-bosluk-analizi.json"
ASG = ROOT / "outputs" / "analiz" / "atama-birlesik.json"
TAX = ROOT / "outputs" / "analiz" / "taxonomy-birlesik.json"
XLSX = ROOT / "chatbot.xlsx"
DST = ROOT / "outputs" / "analiz" / "KB-taslak-eksik-sorular.csv"

WS = re.compile(r"\s+")
MAX_Q = 10

# Botun her session'da tekrarladigi kaliplar; taslak icin degersiz.
BOILER = (
    "iyi günler ben auzef", "merhaba. size hangi konuda", "bunu mu demek istediniz",
    "bu cevap sizin için yeterli", "bir süredir hiçbir şey yazmadınız",
    "sizi ne yazık ki anlayamadım", "yardımcı olabileceğimiz farklı",
    "konuyla ilgili bir talep oluşturmak", "fakültemiz ile ilgili tüm sorularınız",
    "bizimle iletişime geçtiğiniz için", "size nasıl yardımcı", "merhaba, size nasıl",
    "talep oluşturmak istediğiniz bölümü", "sistemde kayıtlı olan tc",
    "talebiniz oluşturulmuştur", "karşılaştığınız sorun için özür",
    "oluşturmak istediğiniz taleple ilgili", "talep oluşturabilmek ya da",
)
TR = str.maketrans({"ı": "i", "İ": "i", "I": "i", "ş": "s", "Ş": "s", "ğ": "g",
                    "Ğ": "g", "ü": "u", "Ü": "u", "ö": "o", "Ö": "o", "ç": "c", "Ç": "c"})
STOP = {"nasil", "nedir", "ne", "mi", "mu", "icin", "bir", "bu", "ile", "var", "nereden"}


def toks(s: str) -> set[str]:
    s = s.translate(TR).casefold()
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode()
    return {w for w in re.findall(r"[a-z0-9]+", s) if len(w) > 2 and w not in STOP}


def assign_tags(questions: list[str], vocab: list[str]) -> dict[str, str]:
    """Her soruya mevcut tag sozlugunden en uygun tag'i sectirir."""
    key = os.environ.get("OPENROUTER_API_KEY", "").strip()
    if not key:
        print("  (OPENROUTER_API_KEY yok — tag'ler bos birakiliyor)")
        return {}
    sys.path.insert(0, str(ROOT / "apps" / "backend"))
    from app.core.config import Settings
    from app.services.openrouter import OpenRouterClient
    from pydantic import BaseModel

    class T(BaseModel):
        soru_no: int
        tag: str

    class Out(BaseModel):
        atamalar: list[T]

    client = OpenRouterClient(
        api_key=key, model="google/gemini-2.5-flash",
        settings=Settings(openrouter_base_url="https://openrouter.ai/api/v1"))
    out = client.complete_structured(
        system=("Sana bir universite chatbot'unun MEVCUT tag sozlugu ve yeni sorular "
                "verilecek. Her soruya sozlukten EN UYGUN tag'i ata. Sozlukte olmayan "
                "bir tag URETME. Hicbiri uymuyorsa bos string yaz."),
        user=("TAG SOZLUGU:\n" + "\n".join(vocab)
              + "\n\nSORULAR:\n" + "\n".join(f"{i}: {q}" for i, q in enumerate(questions, 1))),
        schema={"type": "object", "additionalProperties": False,
                "properties": {"atamalar": {"type": "array", "items": {
                    "type": "object", "additionalProperties": False,
                    "properties": {"soru_no": {"type": "integer"}, "tag": {"type": "string"}},
                    "required": ["soru_no", "tag"]}}},
                "required": ["atamalar"]},
        schema_name="tags", model_type=Out).data.atamalar
    valid = set(vocab)
    return {questions[t.soru_no - 1]: (t.tag if t.tag in valid else "")
            for t in out if 1 <= t.soru_no <= len(questions)}


def main() -> int:
    gaps = [g for g in json.loads(GAP.read_text(encoding="utf-8")) if g["kb_id"] == "yok"]
    gaps = [g for g in gaps if g["adet"] > 0]
    tax = json.loads(TAX.read_text(encoding="utf-8"))
    asg = json.loads(ASG.read_text(encoding="utf-8"))
    kb = load_kb()
    print(f"KB'de karsiligi olmayan, orneklemde gorulen: {len(gaps)} madde")

    qid_of = {q["canonical_question"]: f"q{i}" for i, q in enumerate(tax, 1)}
    want = {qid_of[g["soru"]]: g for g in gaps if g["soru"] in qid_of}
    sessions_of: dict[str, list[dict]] = collections.defaultdict(list)
    for a in asg:
        if a["question_id"] in want:
            sessions_of[a["question_id"]].append(a)
    sid_to_q = {a["session"]: q for q, rows in sessions_of.items() for a in rows}
    print(f"ilgili session: {len(sid_to_q)}")

    # Eski botun bu session'larda verdigi anlamli cevaplar
    print("eski bot cevaplari taraniyor (tek gecis)...", flush=True)
    replies: dict[str, collections.Counter[str]] = collections.defaultdict(collections.Counter)
    wb = load_workbook(XLSX, read_only=True)
    seen = 0
    for name in wb.sheetnames:
        it = wb[name].iter_rows(values_only=True)
        idx = {h: i for i, h in enumerate(next(it))}
        for row in it:
            if row is None:
                continue
            sid = row[idx["SessionId"]]
            if not sid or str(sid) not in sid_to_q:
                continue
            if str(row[idx["Direction"]] or "") != "Bot":
                continue
            t = WS.sub(" ", str(row[idx["MessageTextClean"]] or "")).strip()
            # DIKKAT: "İ".casefold() -> "i̇" (i + birlesen nokta); duz casefold
            # Turkce buyuk I'yi "i"ye indirmez ve filtre sessizce kacirir.
            low = t.translate(TR).casefold()
            if len(t) < 60 or any(low.startswith(b) for b in BOILER):
                continue
            replies[sid_to_q[str(sid)]][t] += 1
            seen += 1
    print(f"  anlamli bot cevabi: {seen}")

    # Tag: mevcut KB'nin 49 tag'lik sozlugunden SEC. Sozluksel eslestirme
    # denendi ve saçmaladi ("Kayit tarihleri" -> UC_DERS_SINAVI); yanlis tag
    # yanlis cevap kadar zararli, o yuzden sozlukten secim yaptiriliyor.
    tag_vocab = sorted({(r.get("tags") or "").strip() for r in kb if (r.get("tags") or "").strip()})
    tag_map = assign_tags([g["soru"] for g in sorted(want.values(), key=lambda x: -x["adet"])],
                          tag_vocab)

    def suggest_tag(q: str) -> str:
        return tag_map.get(q, "")

    header = (["question", "answer", "tags"] + [f"query_{i}" for i in range(1, 21)]
              + ["_hacim", "_sorunlu_yuzde", "_eski_bot_cevabi"])
    rows_out = []
    for qid, g in sorted(want.items(), key=lambda x: -x[1]["adet"]):
        texts = []
        seen_norm: set[str] = set()
        for a in sorted(sessions_of[qid], key=lambda a: a["turns"]):
            for part in a["text"].split(" / "):
                part = part.strip()
                n = " ".join(sorted(toks(part)))
                if 10 <= len(part) <= 120 and n and n not in seen_norm:
                    seen_norm.add(n)
                    texts.append(part)
            if len(texts) >= MAX_Q:
                break
        top_reply = replies[qid].most_common(1)
        rows_out.append(
            [g["soru"], "", suggest_tag(g["soru"])]
            + (texts[:20] + [""] * 20)[:20]
            + [g["adet"], g["sorunlu"], top_reply[0][0][:600] if top_reply else ""]
        )

    with DST.open("w", encoding="utf-8-sig", newline="") as fh:
        w = csv.writer(fh, delimiter=";", quoting=csv.QUOTE_ALL)
        w.writerow(header)
        w.writerows(rows_out)

    withtag = sum(1 for r in rows_out if r[2])
    withreply = sum(1 for r in rows_out if r[-1])
    avgq = sum(sum(1 for x in r[3:23] if x) for r in rows_out) / len(rows_out)
    print(f"\n{len(rows_out)} satir -> {DST}")
    print(f"  tag onerisi bulunan : {withtag}/{len(rows_out)}")
    print(f"  eski bot cevabi olan: {withreply}/{len(rows_out)}")
    print(f"  ortalama gercek ifade: {avgq:.1f}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
