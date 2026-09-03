"""FAZ 3 — atamalardan bes raporu uretir. Deterministik, LLM yok.

Kullanim: python3 scripts/build_reports.py <atama.json> <taksonomi.json>
"""
from __future__ import annotations

import collections
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUT_XLSX = ROOT / "outputs" / "analiz" / "AUZEF-chatbot-analiz-raporu.xlsx"
OUT_MD = ROOT / "outputs" / "analiz" / "RAPOR.md"
HEAD = PatternFill("solid", fgColor="1F3864")
HEADF = Font(color="FFFFFF", bold=True)


def sheet(wb: Workbook, title: str, headers: list[str], rows: list[list]) -> None:
    ws = wb.create_sheet(title[:31])
    ws.append(headers)
    for c in ws[1]:
        c.fill, c.font = HEAD, HEADF
        c.alignment = Alignment(vertical="center", wrap_text=True)
    for r in rows:
        ws.append(r)
    widths = [10, 64, 12, 12, 12, 12, 14, 14]
    for i, w in enumerate(widths[: len(headers)], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def main() -> int:
    asg = json.loads(Path(sys.argv[1]).read_text(encoding="utf-8"))
    tax = json.loads(Path(sys.argv[2]).read_text(encoding="utf-8"))
    label = {f"q{i}": q["canonical_question"] for i, q in enumerate(tax, 1)}
    src = {f"q{i}": ("katalog" if q["source"] == "etiya-qna" else "boşluk")
           for i, q in enumerate(tax, 1)}

    total = len(asg)
    vol = collections.Counter(a["question_id"] for a in asg)
    fail = collections.Counter(a["question_id"] for a in asg if a["rej"] or a["fallback"])
    rej = collections.Counter(a["question_id"] for a in asg if a["rej"])
    fb = collections.Counter(a["question_id"] for a in asg if a["fallback"])
    per_month: dict[str, collections.Counter[str]] = collections.defaultdict(collections.Counter)
    per_chan: dict[str, collections.Counter[str]] = collections.defaultdict(collections.Counter)
    month_tot: collections.Counter[str] = collections.Counter()
    chan_tot: collections.Counter[str] = collections.Counter()
    for a in asg:
        per_month[a["question_id"]][a["month"]] += 1
        per_chan[a["question_id"]][a["channel"]] += 1
        month_tot[a["month"]] += 1
        chan_tot[a["channel"]] += 1
    months = sorted(month_tot)

    wb = Workbook()
    wb.remove(wb.active)

    # 1 — SSS
    rows = []
    for qid, c in vol.most_common():
        if qid == "none":
            continue
        top = per_month[qid].most_common(1)[0]
        rows.append([int(qid[1:]), label[qid], c, round(c / total * 100, 2),
                     round(fail[qid] / c * 100), src[qid], top[0],
                     round(top[1] / c * 100)])
    sheet(wb, "1-SSS (hacim)",
          ["#", "Soru", "Adet", "Pay %", "Sorunlu %", "Kaynak", "Zirve ay", "Zirve payı %"], rows)

    # 3 — KB duzeltme: cevabi var, reddediliyor. Hacim esigi 50.
    fixes = sorted(((q, c) for q, c in vol.items() if q != "none" and c >= 50),
                   key=lambda x: -(fail[x[0]] / x[1]))
    sheet(wb, "3-KB duzeltme (ret orani)",
          ["#", "Soru", "Adet", "Sorunlu %", "Ret", "Anlamadi", "Kaynak"],
          [[int(q[1:]), label[q], c, round(fail[q] / c * 100), rej[q], fb[q], src[q]]
           for q, c in fixes])

    # 4 — Mevsimsellik
    rows = [["#", "Soru", "Toplam"] + months]
    body = []
    for qid, c in vol.most_common(60):
        if qid == "none":
            continue
        body.append([int(qid[1:]), label[qid], c] + [per_month[qid].get(m, 0) for m in months])
    sheet(wb, "4-Mevsimsellik", rows[0], body)

    # 5 — Kanal
    chans = sorted(chan_tot)
    body = []
    for qid, c in vol.most_common(60):
        if qid == "none":
            continue
        row = [int(qid[1:]), label[qid], c]
        for ch in chans:
            n = per_chan[qid].get(ch, 0)
            row += [n, round(n / c * 100)]
        body.append(row)
    hdr = ["#", "Soru", "Toplam"]
    for ch in chans:
        hdr += [ch, f"{ch} %"]
    sheet(wb, "5-Kanal", hdr, body)

    # 2 — KB yeni makale: «hicbiri» kovasi
    none = [a for a in asg if a["question_id"] == "none"]
    nf = sum(1 for a in none if a["rej"] or a["fallback"])
    step = max(1, len(none) // 200)
    sheet(wb, "2-KB yeni (hicbiri)",
          ["Ay", "Kanal", "Sorunlu", "Konusma"],
          [[a["month"], a["channel"], "evet" if (a["rej"] or a["fallback"]) else "hayır",
            a["text"][:500]] for a in none[::step][:200]])

    wb.save(OUT_XLSX)

    md = [f"# AUZEF Chatbot Analiz Raporu", "",
          f"Kaynak: bir yıllık döküm (2025-07-10 → 2026-07-10), 6.405.948 satır",
          f"Yazılı session: 289.567 · Örneklem: {total} (aylara orantılı, sistematik)",
          f"Taksonomi: {len(tax)} madde ({sum(1 for v in src.values() if v=='katalog')} "
          f"eski bot kataloğundan, {sum(1 for v in src.values() if v=='boşluk')} boşluk analizinden)", "",
          "## Özet", "",
          f"- Sınıflandırılan: {total - vol['none']} (%{(total-vol['none'])/total*100:.1f})",
          f"- Taksonomiye oturmayan: {vol['none']} (%{vol['none']/total*100:.1f}), "
          f"içinin %{nf/len(none)*100:.0f}'i sorunlu",
          f"- Sorunlu session (ret veya anlamadı): "
          f"{sum(fail.values())} (%{sum(fail.values())/total*100:.1f})", "",
          "## En çok sorulan 15", "",
          "| # | Soru | Adet | Pay | Sorunlu |", "|---|---|---|---|---|"]
    for qid, c in vol.most_common(16):
        if qid == "none":
            continue
        md.append(f"| {qid[1:]} | {label[qid]} | {c} | %{c/total*100:.1f} | "
                  f"%{fail[qid]/c*100:.0f} |")
    md += ["", "## En sorunlu 15 (hacmi 50+ olanlar arasında)", "",
           "| # | Soru | Adet | Sorunlu |", "|---|---|---|---|"]
    for q, c in fixes[:15]:
        md.append(f"| {q[1:]} | {label[q]} | {c} | %{fail[q]/c*100:.0f} |")
    md += ["", "## Aylık örneklem hacmi", ""]
    for m in months:
        md.append(f"- {m}: {month_tot[m]}")
    OUT_MD.write_text("\n".join(md), encoding="utf-8")

    print(f"xlsx -> {OUT_XLSX}")
    print(f"md   -> {OUT_MD}")
    print(f"\ntoplam {total} | siniflandirilan {total - vol['none']} "
          f"(%{(total-vol['none'])/total*100:.1f}) | hicbiri {vol['none']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
