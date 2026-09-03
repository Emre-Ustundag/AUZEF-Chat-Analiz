from __future__ import annotations

import argparse
import csv
import json
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path


ROOT = Path(__file__).resolve().parents[2]
BACKEND = ROOT / "apps" / "backend"
sys.path.insert(0, str(BACKEND))

from app.core.config import Settings  # noqa: E402
from app.core.catalog import DEFAULT_MODEL, DEFAULT_PROMPT_VERSION  # noqa: E402
from app.pipeline.cost import estimate_cost, estimate_profile_cost  # noqa: E402
from app.pipeline.preprocess import (  # noqa: E402
    COURTESY_ONLY,
    is_only_masks,
    is_system_message,
    normalize,
    preprocess,
)
from app.prompts.faq_analysis import get_prompt  # noqa: E402
from app.services.redaction import redact_pii  # noqa: E402
from app.services.xlsx import profile_xlsx, validate_xlsx  # noqa: E402
from openpyxl import Workbook, load_workbook  # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill  # noqa: E402


SAMPLE_SIZES = (120, 600, 2400)
SHEET_NAME = "Mesajlar"
TIME_COLUMN = "message_time_tr"
TEXT_COLUMN = "message_text_clean"


@dataclass(frozen=True)
class Candidate:
    timestamp: datetime
    text: str
    normalized: str
    source_row: int


def parse_timestamp(value: str) -> datetime | None:
    text = value.strip()
    if not text:
        return None

    normalized = text.removesuffix("Z").replace("T", " ")
    try:
        return datetime.fromisoformat(normalized)
    except ValueError:
        pass

    for pattern in (
        "%d.%m.%Y %H:%M:%S",
        "%d.%m.%Y %H:%M",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
    ):
        try:
            return datetime.strptime(normalized, pattern)
        except ValueError:
            continue
    return None


def systematic_pick(items: list[Candidate], size: int) -> list[Candidate]:
    if size > len(items):
        raise ValueError(f"{size} kayıt istendi fakat yalnızca {len(items)} aday var")
    if size == 1:
        return [items[len(items) // 2]]

    last = len(items) - 1
    indexes = [round(index * last / (size - 1)) for index in range(size)]
    if len(set(indexes)) != size:
        raise RuntimeError("Sistematik örnekleme benzersiz indis üretemedi")
    return [items[index] for index in indexes]


def collect_candidates(source: Path, settings: Settings) -> tuple[list[Candidate], dict[str, int]]:
    csv.field_size_limit(sys.maxsize)
    candidates_by_normalized: dict[str, Candidate] = {}
    stats = {
        "source_rows": 0,
        "missing_text": 0,
        "invalid_time": 0,
        "discarded_preprocess": 0,
        "normalized_duplicates": 0,
    }

    with source.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        required = {TIME_COLUMN, TEXT_COLUMN}
        missing = required.difference(reader.fieldnames or [])
        if missing:
            raise ValueError(f"Eksik CSV kolonları: {sorted(missing)}")

        for source_row, row in enumerate(reader, start=2):
            stats["source_rows"] += 1
            raw = row.get(TEXT_COLUMN)
            if raw is None:
                stats["missing_text"] += 1
                continue

            text = " ".join(raw.split()).strip()
            if (
                not text
                or len(text) < settings.preprocess_min_message_length
                or is_system_message(text)
                or normalize(text) in COURTESY_ONLY
            ):
                stats["discarded_preprocess"] += 1
                continue

            redacted = redact_pii(text)
            normalized = normalize(redacted)
            if not normalized or is_only_masks(normalized):
                stats["discarded_preprocess"] += 1
                continue

            timestamp = parse_timestamp(row.get(TIME_COLUMN, ""))
            if timestamp is None:
                stats["invalid_time"] += 1
                continue

            candidate = Candidate(
                timestamp=timestamp,
                text=text,
                normalized=normalized,
                source_row=source_row,
            )
            previous = candidates_by_normalized.get(normalized)
            if previous is None:
                candidates_by_normalized[normalized] = candidate
            else:
                stats["normalized_duplicates"] += 1
                if (candidate.timestamp, candidate.source_row) < (
                    previous.timestamp,
                    previous.source_row,
                ):
                    candidates_by_normalized[normalized] = candidate

    candidates = sorted(
        candidates_by_normalized.values(),
        key=lambda item: (item.timestamp, item.source_row, item.normalized),
    )
    stats["eligible_unique"] = len(candidates)
    return candidates, stats


def write_workbook(path: Path, candidates: list[Candidate]) -> None:
    workbook = Workbook(write_only=False)
    sheet = workbook.active
    sheet.title = SHEET_NAME
    sheet.append([TIME_COLUMN, TEXT_COLUMN])

    for candidate in candidates:
        sheet.append([candidate.timestamp, candidate.text])

    header = sheet[1]
    fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    for cell in header:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="left", vertical="center")

    for cell in sheet["A"]:
        if cell.row > 1:
            cell.number_format = "yyyy-mm-dd hh:mm:ss"

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:B{sheet.max_row}"
    sheet.column_dimensions["A"].width = 22
    sheet.column_dimensions["B"].width = 100
    sheet.row_dimensions[1].height = 24
    workbook.save(path)


def verify_workbook(path: Path, expected_size: int) -> dict[str, object]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if workbook.sheetnames != [SHEET_NAME]:
        raise AssertionError(f"Beklenmeyen sayfalar: {workbook.sheetnames}")

    sheet = workbook[SHEET_NAME]
    rows = sheet.iter_rows(values_only=True)
    header = next(rows)
    if header != (TIME_COLUMN, TEXT_COLUMN):
        raise AssertionError(f"Beklenmeyen başlık: {header}")

    settings = Settings(_env_file=None)
    normalized_values: set[str] = set()
    timestamps: list[datetime] = []
    raw_values: list[str] = []
    count = 0
    for timestamp, raw in rows:
        count += 1
        if not isinstance(timestamp, datetime):
            raise AssertionError(f"{count + 1}. satırda tarih hücresi datetime değil")
        if not isinstance(raw, str) or not raw.strip():
            raise AssertionError(f"{count + 1}. satırda mesaj boş")

        text = " ".join(raw.split()).strip()
        if len(text) < settings.preprocess_min_message_length or is_system_message(text):
            raise AssertionError(f"{count + 1}. satır ön işlemede elenir")
        normalized = normalize(redact_pii(text))
        if not normalized or normalized in COURTESY_ONLY or is_only_masks(normalized):
            raise AssertionError(f"{count + 1}. satır ön işlemede elenir")
        normalized_values.add(normalized)
        timestamps.append(timestamp)
        raw_values.append(raw)

    workbook.close()
    if count != expected_size:
        raise AssertionError(f"{path.name}: {expected_size} yerine {count} kayıt")
    if len(normalized_values) != expected_size:
        raise AssertionError(
            f"{path.name}: {expected_size} yerine {len(normalized_values)} benzersiz kayıt"
        )

    uncompressed_bytes = validate_xlsx(path, settings)
    app_profile = profile_xlsx(path, settings)
    app_sheet = app_profile["sheets"][0]
    app_column = next(
        column for column in app_sheet["columns"] if column["name"] == TEXT_COLUMN
    )
    if app_sheet["row_count"] != expected_size:
        raise AssertionError("Uygulama profili beklenen satır sayısını vermedi")
    if app_column["unique_count"] != expected_size:
        raise AssertionError("Uygulama profili beklenen benzersiz sayısını vermedi")

    estimated_cost_usd = estimate_profile_cost(
        app_column["unique_count"],
        app_column["avg_length"],
        DEFAULT_MODEL,
        settings=settings,
        prompt=get_prompt(DEFAULT_PROMPT_VERSION),
    )
    worker_preprocess = preprocess(raw_values, settings)
    worker_estimate = estimate_cost(
        worker_preprocess.groups,
        DEFAULT_MODEL,
        max_cost_usd=1.0,
        settings=settings,
        prompt=get_prompt(DEFAULT_PROMPT_VERSION),
    )

    return {
        "file": path.name,
        "rows": count,
        "unique_after_preprocess": len(normalized_values),
        "min_time": min(timestamps).isoformat(sep=" "),
        "max_time": max(timestamps).isoformat(sep=" "),
        "size_bytes": path.stat().st_size,
        "uncompressed_bytes": uncompressed_bytes,
        "app_profile_avg_length": app_column["avg_length"],
        "profile_estimated_cost_usd": estimated_cost_usd,
        "worker_estimated_prompt_tokens": worker_estimate.estimated_prompt_tokens,
        "worker_estimated_completion_tokens": worker_estimate.estimated_completion_tokens,
        "worker_estimated_cost_usd": worker_estimate.estimated_cost_usd,
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("source", type=Path)
    parser.add_argument("output_dir", type=Path)
    args = parser.parse_args()

    settings = Settings(_env_file=None)
    candidates, source_stats = collect_candidates(args.source, settings)
    master = systematic_pick(candidates, SAMPLE_SIZES[-1])
    medium = systematic_pick(master, SAMPLE_SIZES[1])
    small = systematic_pick(medium, SAMPLE_SIZES[0])
    samples = {120: small, 600: medium, 2400: master}

    args.output_dir.mkdir(parents=True, exist_ok=True)
    verifications: list[dict[str, object]] = []
    for size, sample in samples.items():
        path = args.output_dir / f"kalibrasyon-{size}.xlsx"
        write_workbook(path, sample)
        verifications.append(verify_workbook(path, size))

    normalized_sets = {
        size: {candidate.normalized for candidate in sample}
        for size, sample in samples.items()
    }
    if not normalized_sets[120] <= normalized_sets[600] <= normalized_sets[2400]:
        raise AssertionError("Örnek dosyalar birbirinin alt kümesi değil")

    summary = {
        "source": args.source.name,
        "preprocess_min_message_length": settings.preprocess_min_message_length,
        "source_stats": source_stats,
        "nested_samples": True,
        "workbooks": verifications,
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
